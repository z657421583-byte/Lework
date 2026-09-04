import io
import json
import sys
import tempfile
import time
import unittest
from contextlib import redirect_stdout
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from scripts.payroll_calculator import (
    attendance_records,
    attendance_records_from_paths,
    calculate,
    load_historical_payroll,
    main,
    month_statutory_holidays,
    note_skipped_generated_history,
    parse_construction_formula,
    parse_performance_standard,
    read_rows,
    source_hints,
    fuzzy_match_review_note,
    write_workbook,
)


class PayrollCalculatorTest(unittest.TestCase):
    def make_book(self, directory, name, headers, rows):
        path = Path(directory) / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path

    def test_aliases_partial_baseline_and_rules(self):
        with tempfile.TemporaryDirectory() as directory:
            roster = self.make_book(
                directory, "roster.xlsx", ["姓名", "项目", "人员类别", "状态"],
                [["张三", "A项目", "外聘", "正常"], ["李四", "A项目", "外聘", "在册"]],
            )
            history = self.make_book(
                directory, "history.xlsx",
                ["姓名", "项目", "人员类别", "基本工资", "绩效工资", "工龄工资",
                 "职称工资", "施工补贴", "1-3月话补", "降温费", "交通补助", "双休日加班", "应发工资"],
                [["张三", "A项目", "外聘", 10000, 3000, 200, 100, 50, 300, 200, 100, 920, 0],
                 ["李四", "A项目", "外聘", 9000, 2000, 0, 0, 40, 200, 0, 0, 0, 0]],
            )
            rows = calculate(
                read_rows(roster), read_rows(history),
                [{"name": "张三", "project": "A项目", "category": "外聘",
                  "actual_work_days": 23, "personal_leave_days": 2}],
                "2026-06", 22, 2,
            )
            detail = rows["payroll_detail"][0]
            self.assertEqual(detail["绩效工资"], 3000 - 3000 / 21.75 * 2)
            self.assertEqual(detail["施工补贴"], 0)
            self.assertEqual(detail["加班天数"], 2)
            self.assertEqual(rows["attendance"][0]["来源"], "当月考勤表")
            self.assertTrue(any("施工补贴只有月金额" in item["说明"]
                                for item in rows["review_exceptions"]))
            self.assertTrue(any(item["类型"] == "可忽略" for item in rows["review_exceptions"]))
            self.assertTrue(any(item.get("姓名") == "张三" for item in rows["review_exceptions"]))

            output = Path(directory) / "result.xlsx"
            write_workbook(rows, output)
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["工资核算明细", "工资基准", "考勤汇总", "历史工资对比", "人工复核事项"],
                )
            finally:
                workbook.close()

    def test_missing_workdays_does_not_block_other_items(self):
        rows = calculate(
            [], [{"name": "王五", "project": "B", "category": "外聘",
                  "position_salary": 5000, "performance": 1000, "overtime_amount": 460}],
            [{"name": "王五", "project": "B", "category": "外聘",
              "actual_work_days": 20}],
            "2026-05", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["岗位工资"], 5000)
        self.assertEqual(rows["payroll_detail"][0]["加班费"], 460)
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 1)

    def test_missing_attendance_is_not_written_as_zero(self):
        rows = calculate(
            [], [{"name": "缺考勤", "project": "B", "category": "外聘",
                  "position_salary": 5000, "performance": 1000}],
            [{"name": "缺考勤", "project": "B", "category": "外聘"}],
            "2026-06", 21, 4,
        )
        detail = rows["payroll_detail"][0]
        self.assertIsNone(detail["实际出勤"])
        self.assertTrue(any("缺少实际出勤" in item["说明"] for item in rows["review_exceptions"]))

    def test_historical_zero_overtime_is_not_calculated(self):
        rows = calculate(
            [],
            [{"name": "一口价", "project": "B", "category": "外聘",
              "position_salary": 7500, "performance": 0, "overtime_amount": 0}],
            [{"name": "一口价", "project": "B", "category": "外聘",
              "actual_work_days": 25}],
            "2026-06", None, None,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["加班天数"], 0)
        self.assertEqual(detail["加班费"], 0)
        self.assertEqual(detail["应发工资"], 7500)
        self.assertIsNone(rows["baseline"][0]["加班标准"])

    def test_missing_historical_overtime_is_not_derived_from_salary(self):
        rows = calculate(
            [],
            [{"name": "一口价", "project": "B", "category": "外聘",
              "position_salary": 7000}],
            [{"name": "一口价", "project": "B", "category": "外聘",
              "actual_work_days": 25}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["加班费"], 0)
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 0)

    def test_overtime_rate_column_is_not_eligibility(self):
        rows = calculate(
            [],
            [{"name": "一口价", "project": "B", "category": "外聘",
              "position_salary": 2400, "performance": 3100,
              "overtime_standard": 220, "overtime_count": 0, "overtime_amount": 0}],
            [{"name": "一口价", "project": "B", "category": "外聘",
              "actual_work_days": 27}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 0)
        self.assertEqual(rows["payroll_detail"][0]["加班费"], 0)
        self.assertIsNone(rows["baseline"][0]["加班标准"])

    def test_grouped_overtime_headers_read_amount_not_rate(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "姓名"
            sheet["B1"] = "基本工资"
            sheet["C1"] = "绩效工资"
            sheet["D1"] = "工作天数"
            sheet.merge_cells("E1:G1")
            sheet["E1"] = "双休日加班"
            sheet["E2"] = "标准"
            sheet["F2"] = "个数"
            sheet["G2"] = "金额"
            sheet["A3"] = "郭胜松"
            sheet["B3"] = 2400
            sheet["C3"] = 3100
            sheet["D3"] = 27
            sheet["E3"] = "=ROUND(B3/21.75,0)*2"
            sheet["G3"] = "=E3*F3"
            sheet["A4"] = "张海峰"
            sheet["B4"] = 2000
            sheet["C4"] = 2000
            sheet["D4"] = 26
            sheet["E4"] = "=ROUND(B4/21.75,0)*2"
            sheet["F4"] = 4
            sheet["G4"] = "=E4*F4"
            workbook.save(path)
            loaded = {row["name"]: row for row in read_rows(path)}
            self.assertEqual(loaded["郭胜松"]["overtime_amount"], 0)
            self.assertEqual(loaded["郭胜松"]["overtime_standard"], 220)
            self.assertGreater(loaded["张海峰"]["overtime_amount"], 0)
            rows = calculate(
                [], read_rows(path),
                [{"name": "郭胜松", "actual_work_days": 27},
                 {"name": "张海峰", "actual_work_days": 26}],
                "2026-06", None, None,
            )
            details = {row["姓名"]: row for row in rows["payroll_detail"]}
            self.assertEqual(details["郭胜松"]["加班费"], 0)
            self.assertGreater(details["张海峰"]["加班费"], 0)

    def test_merged_attendance_over_calendar_month_is_reviewed(self):
        rows = calculate(
            [],
            [{"name": "沈军鹏", "project": "戏剧学院", "category": "外聘",
              "position_salary": 2700, "performance": 3800, "overtime_amount": 248,
              "work_days": 27, "_source_month": "2026-06"}],
            [
                {"name": "沈军鹏", "project": "杨职院", "actual_work_days": 11},
                {"name": "沈军鹏", "project": "戏剧学院", "actual_work_days": 26},
            ],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["实际出勤"], 37)
        notes = "；".join(
            item["说明"] for item in rows["review_exceptions"] if item.get("姓名") == "沈军鹏"
        )
        self.assertIn("超过", notes)
        self.assertIn("工作天数", notes)

    def test_shuangxiuri_header_maps_to_overtime_amount(self):
        with tempfile.TemporaryDirectory() as directory:
            history = self.make_book(
                directory, "history.xlsx",
                ["姓名", "项目", "人员类别", "基本工资", "绩效工资", "双休日加班"],
                [["张三", "A项目", "外聘", 5000, 1000, 460]],
            )
            rows = calculate(
                [], read_rows(history),
                [{"name": "张三", "project": "A项目", "category": "外聘",
                  "actual_work_days": 22}],
                "2026-06", None, None,
            )
            self.assertEqual(rows["payroll_detail"][0]["加班天数"], 1)
            self.assertEqual(rows["payroll_detail"][0]["加班费"], 460)

    def test_construction_total_is_normalized_to_daily_standard(self):
        rows = calculate(
            [],
            [{"name": "赵六", "项目": "C", "category": "外聘",
              "work_days": 26, "position_salary": 5000,
              "performance": 1000, "construction": 780}],
            [{"name": "赵六", "project": "C", "category": "外聘",
              "actual_work_days": 23}],
            "2026-06", 22, 8,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["施工补贴"], 690)

    def test_hot_subsidy_requires_individual_history_and_explicit_overtime_is_counted(self):
        rows = calculate(
            [],
            [{"name": "甲", "项目": "D", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "hot": 600, "construction_day": 30, "overtime_amount": 460},
             {"name": "乙", "项目": "D", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "hot": 0, "construction_day": 30, "overtime_amount": 460}],
            [{"name": "甲", "project": "D", "category": "外聘",
              "actual_work_days": 22, "加班天数": 2},
             {"name": "乙", "project": "D", "category": "外聘",
              "actual_work_days": 22}],
            "2026-06", 22, 8,
        )
        details = {row["姓名"]: row for row in rows["payroll_detail"]}
        baselines = {row["姓名"]: row for row in rows["baseline"]}
        self.assertEqual(baselines["甲"]["高温补贴"], 600)
        self.assertEqual(baselines["乙"]["高温补贴"], 0)
        self.assertEqual(details["甲"]["加班天数"], 1)
        self.assertEqual(details["甲"]["加班费"], 460)

    def test_quarter_phone_is_paid_in_quarter_end_month(self):
        rows = calculate(
            [],
            [{"name": "钱七", "项目": "E", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "phone_1_3": 300}],
            [{"name": "钱七", "project": "E", "category": "外聘",
              "actual_work_days": 22}],
            "2026-03", 22, 0,
        )
        self.assertEqual(rows["payroll_detail"][0]["话费补贴"], 300)

    def test_overtime_counts_attendance_on_single_rest_day(self):
        rows = calculate(
            [],
            [{"name": "孙八", "项目": "F", "category": "外聘",
              "position_salary": 5000, "performance": 1000, "overtime_amount": 460}],
            [{"name": "孙八", "project": "F", "category": "外聘",
              "actual_work_days": 27,
              "daily_marks": {
                  "2026-06-06": "出勤",  # Saturday is a planned workday for single rest.
                  "2026-06-07": "出勤",  # Both weekend days worked: two overtime days.
              }}],
            "2026-06", 26, 4, "single",
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 4)

    def test_overtime_counts_one_day_for_each_worked_weekend_without_schedule(self):
        rows = calculate(
            [], [{"name": "周末", "project": "A项目", "category": "外包",
                  "position_salary": 3800, "performance": 1000, "overtime_amount": 1050}],
            [{"name": "周末", "project": "A项目", "category": "外包",
              "actual_work_days": 24, "weekend_attendance_dates": [
                  "2026-06-06", "2026-06-14", "2026-06-20", "2026-06-28",
              ]}],
            "2026-06", None, None,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["加班天数"], 3)
        self.assertEqual(detail["加班费"], 1050)

    def test_june_uses_verified_overtime_public_rule(self):
        rows = calculate(
            [], [{"name": "赵六", "project": "A项目", "category": "外包",
                  "position_salary": 3800, "performance": 1000, "overtime_amount": 1400}],
            [{"name": "赵六", "project": "A项目", "category": "外包",
              "actual_work_days": 26, "overtime_days": 4}],
            "2026-06", 26, 4, "single",
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["加班天数"], 4)
        self.assertEqual(detail["加班费"], 1400)

    def test_june_holiday_reduces_base_workdays_before_overtime_cap(self):
        rows = calculate(
            [], [{"name": "节日", "project": "A项目", "category": "外包",
                  "position_salary": 3800, "performance": 1000, "overtime_amount": 1400}],
            [{"name": "节日", "project": "A项目", "category": "外包",
              "actual_work_days": 25, "holiday_dates": [
                  "2026-06-19", "2026-06-20", "2026-06-21",
              ]}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 4)

    def test_unique_ocr_name_correction_is_marked_for_review(self):
        rows = calculate(
            [], [{"name": "陈丽", "project": "A项目", "category": "外包",
                  "position_salary": 3000}],
            [{"name": "陈利", "project": "A项目", "category": "外包",
              "actual_work_days": 21}],
            "2026-06", 21, 4, "single",
        )
        self.assertEqual(rows["payroll_detail"][0]["项目"], "A项目")
        self.assertTrue(any("模糊匹配" in item["说明"] for item in rows["review_exceptions"]))

    def test_name_ocr_correction_is_reported_once(self):
        rows = calculate(
            [{"name": "周罡", "project": "A项目", "category": "外包"}],
            [{"name": "周罡", "project": "A项目", "category": "外包",
              "position_salary": 3000}],
            [{"name": "周翌", "project": "A项目", "category": "外包",
              "actual_work_days": 21}],
            "2026-06", 21, 4,
        )
        notes = [
            item["说明"] for item in rows["review_exceptions"]
            if item["姓名"] == "周翌" and "模糊匹配" in item["说明"]
        ]
        self.assertEqual(notes, ["姓名单字 OCR 易混，已模糊匹配，需人工复核是否同一人"])

    def test_unique_near_name_still_calculates_and_asks_review(self):
        rows = calculate(
            [{"name": "苏亚南", "category": "人事代理", "_source_file": "人员花名册.xlsx", "_row": 55}],
            [{"name": "苏亚南", "category": "人事代理", "position_salary": 3000, "performance": 1000,
              "_source_file": "2026年6月份人事代理.xlsx", "_row": 55}],
            [{"name": "苏亚楠", "actual_work_days": 21}],
            "2026-06", None, None,
        )
        self.assertEqual(len(rows["payroll_detail"]), 1)
        self.assertEqual(rows["payroll_detail"][0]["岗位工资"], 3000)
        notes = "；".join(item["说明"] for item in rows["review_exceptions"])
        self.assertIn("模糊匹配", notes)
        self.assertNotIn("未计算", notes)
        self.assertNotIn("歧义", notes)
        self.assertNotIn("姓名姓名", notes)

    def test_fuzzy_review_note_does_not_repeat_name_prefix(self):
        self.assertEqual(
            fuzzy_match_review_note("编辑相似度 80%"),
            "姓名编辑相似度 80%，已模糊匹配，需人工复核是否同一人",
        )
        self.assertNotIn("姓名姓名", fuzzy_match_review_note("姓名编辑相似度 80%"))

    def test_other_month_history_workdays_are_not_compared(self):
        rows = calculate(
            [],
            [{"name": "甲", "category": "外聘", "position_salary": 3000, "performance": 1000,
              "work_days": 21, "_source_month": "2026-05"}],
            [{"name": "甲", "category": "外聘", "actual_work_days": 26}],
            "2026-06", None, None,
        )
        notes = "；".join(item["说明"] for item in rows["review_exceptions"] if item.get("姓名") == "甲")
        self.assertNotIn("工作天数", notes)

    def test_roster_does_not_override_attendance_project(self):
        rows = calculate(
            [{"name": "甲", "project": "A项目", "category": "外聘", "status": "正常"}],
            [{"name": "甲", "project": "A项目", "category": "外聘",
              "position_salary": 3000, "performance": 500}],
            [{"name": "甲", "project": "A项目西区", "category": "外聘",
              "actual_work_days": 21}],
            "2026-06", 21, 4,
            attendance_project="A项目西区",
        )
        self.assertEqual(rows["payroll_detail"][0]["项目"], "A项目西区")

    def test_job_title_misread_as_project_does_not_block_matching(self):
        rows = calculate(
            [{"name": "吴工", "project": "A项目", "category": "外包", "status": "正常"}],
            [{"name": "吴工", "project": "A项目", "category": "外包",
              "position_salary": 2300, "performance": 2500}],
            [{"name": "吴工", "project": "施工员", "actual_work_days": 26}],
            "2026-06", 21, 4, attendance_project="A项目",
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["项目"], "A项目")
        self.assertEqual(detail["岗位工资"], 2300)

    def test_unknown_vision_placeholder_is_not_a_payroll_person(self):
        rows = calculate(
            [], [{"name": "未知人员", "project": "A项目", "category": "外聘",
                  "position_salary": 3000}],
            [{"name": "无法识别", "project": "施工员", "actual_work_days": 20}],
            "2026-06", 21, 4, attendance_project="A项目",
        )
        self.assertEqual(rows["payroll_detail"], [])

    def test_unmatched_person_without_history_is_not_paid(self):
        rows = calculate(
            [],
            [],
            [{"name": "周八", "project": "A项目", "category": "甲司外包",
              "actual_work_days": 21}],
            "2026-06", None, None, attendance_project="A项目",
        )
        self.assertEqual(rows["payroll_detail"], [])
        self.assertTrue(any("未匹配历史工资" in item["说明"] for item in rows["review_exceptions"]))

    def test_ignored_review_always_uses_enrolled_category(self):
        rows = calculate(
            [{"name": "钱在册", "project": "A项目", "category": "项目经理", "status": "在册"}],
            [], [{"name": "钱在册", "project": "A项目", "category": "项目经理",
                   "actual_work_days": 21}],
            "2026-06", 21, 4,
        )
        review = rows["review_exceptions"][0]
        self.assertEqual(review["人员类别"], "在册")

    def test_department_label_does_not_block_employment_category_match(self):
        rows = calculate(
            [], [{"name": "孙山", "project": "A项目", "category": "外包",
                  "position_salary": 3000, "performance": 500}],
            [{"name": "孙山", "project": "A项目", "category": "项目管理人员",
              "actual_work_days": 21}],
            "2026-06", 21, 4, "single",
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["人员类别"], "外包")
        self.assertEqual(detail["岗位工资"], 3000)

    def test_project_suffix_and_missing_category_still_match_unique_person(self):
        rows = calculate(
            [{"name": "周九", "project": "A项目", "category": "外聘", "status": "正常"}],
            [{"name": "周九", "project": "A项目", "category": "外聘",
              "position_salary": 5000, "performance": 1000,
              "construction_day": 20}],
            [{"name": "周九", "project": "A项目西区",
              "actual_work_days": 20}],
            "2026-06", 22, 8,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["岗位工资"], 5000)
        self.assertEqual(detail["施工补贴"], 400)
        self.assertEqual(detail["人员类别"], "外聘")
        self.assertEqual(detail["计算状态"], "需复核")

    def test_chinese_attendance_days_are_read(self):
        rows = calculate(
            [{"name": "吴十", "project": "G", "category": "外聘", "status": "正常"}],
            [{"name": "吴十", "project": "G", "category": "外聘",
              "position_salary": 5000, "performance": 1000, "construction_day": 20}],
            [{"name": "吴十", "project": "G", "实际出勤天数": 23}],
            "2026-06", 22, 8,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["实际出勤"], 23)
        self.assertEqual(detail["施工补贴"], 460)

    def test_vision_actual_attendance_and_top_level_project_are_used(self):
        with tempfile.TemporaryDirectory() as directory:
            attendance_path = Path(directory) / "attendance.json"
            attendance_path.write_text(
                '{"month":"2026-06","project":"A项目西区","records":'
                '[{"name":"赵六","actual_attendance":26}]}',
                encoding="utf-8",
            )
            month, project, attendance = attendance_records(attendance_path)
            rows = calculate(
                [{"name": "赵六", "project": "A项目", "category": "外包", "status": "正常"}],
                [{"name": "赵六", "project": "A项目", "category": "外包",
                  "position_salary": 5000, "performance": 1000, "construction_day": 20}],
                attendance, month, 22, 8, attendance_project=project,
            )
            detail = rows["payroll_detail"][0]
            review = rows["review_exceptions"][0]
            self.assertEqual(detail["实际出勤"], 26)
            self.assertEqual(detail["施工补贴"], 520)
            self.assertEqual(detail["项目"], "A项目西区")
            self.assertEqual(detail["人员类别"], "外包")
            self.assertIn("人员类别", review)

    def test_parenthetical_name_annotation_matches_unique_roster_and_history(self):
        rows = calculate(
            [{"name": "李雷（大）", "project": "B项目", "category": "外包", "status": "正常"}],
            [{"name": "李雷", "project": "B项目", "category": "外包",
              "position_salary": 3200, "performance": 4200, "seniority": 240,
              "work_days": 26, "construction": 780}],
            [{"name": "李雷", "project": "B项目", "category": "外包",
              "actual_work_days": 26}],
            "2026-06", None, None,
        )
        detail = rows["payroll_detail"][0]
        self.assertEqual(detail["工龄工资"], 240)
        self.assertEqual(detail["施工补贴"], 780)

    def test_fuzzy_name_candidate_is_not_used_when_tied(self):
        rows = calculate(
            [
                {"name": "王志刚", "project": "B项目", "category": "外包", "status": "正常"},
                {"name": "王志钢", "project": "B项目", "category": "外包", "status": "正常"},
            ],
            [], [{"name": "王志强", "project": "B项目", "category": "外包", "actual_work_days": 21}],
            "2026-06", None, None,
        )
        self.assertTrue(any("模糊匹配存在歧义" in item["说明"] for item in rows["review_exceptions"]))
        self.assertEqual(rows["payroll_detail"], [])

    def test_quarter_phone_is_not_paid_twice_when_already_paid(self):
        history = [
            {"name": "钱七", "project": "E", "category": "外聘", "position_salary": 5000,
             "performance": 1000, "phone_4_6": 300, "_source_month": "2026-05"},
            {"name": "钱七", "project": "E", "category": "外聘", "position_salary": 5000,
             "performance": 1000, "phone_4_6": 300, "_source_month": "2026-06"},
        ]
        rows = calculate(
            [], history,
            [{"name": "钱七", "project": "E", "category": "外聘", "actual_work_days": 21}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["话费补贴"], 0)
        self.assertTrue(any("不重复计入" in item["说明"] for item in rows["review_exceptions"]))

    def test_calendar_fallback_and_workbook_freeze_panes(self):
        rows = calculate(
            [], [{"name": "节日", "project": "B项目", "category": "外聘",
                  "position_salary": 3800, "performance": 1000, "overtime_amount": 1400}],
            [{"name": "节日", "project": "B项目", "category": "外聘", "actual_work_days": 26}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 4)
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            write_workbook(rows, output)
            workbook = load_workbook(output, read_only=False)
            try:
                self.assertTrue(all(sheet.freeze_panes == "A2" for sheet in workbook.worksheets))
            finally:
                workbook.close()

    def test_june_statutory_holidays_include_dragon_boat_weekend_days(self):
        self.assertEqual(
            month_statutory_holidays("2026-06", ["2026-06-06", "2026-06-07"]),
            ["2026-06-19", "2026-06-20", "2026-06-21"],
        )

    def test_multiple_roster_and_attendance_files_are_merged(self):
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "a.json"
            second = Path(directory) / "b.json"
            first.write_text(
                '{"month":"2026-06","records":[{"name":"甲","project":"A","category":"外聘","actual_work_days":21}]}',
                encoding="utf-8",
            )
            second.write_text(
                '{"month":"2026-06","records":[{"name":"乙","project":"B","category":"外聘","actual_work_days":21}]}',
                encoding="utf-8",
            )
            month, _, records = attendance_records_from_paths([first, second])
            self.assertEqual(month, "2026-06")
            self.assertEqual({row["name"] for row in records}, {"甲", "乙"})
            roster_a = self.make_book(
                directory, "roster-a.xlsx", ["姓名", "项目", "人员类别"],
                [["甲", "A", "外聘"]],
            )
            roster_b = self.make_book(
                directory, "roster-b.xlsx", ["姓名", "项目", "人员类别"],
                [["乙", "B", "外聘"]],
            )
            history = self.make_book(
                directory, "history.xlsx",
                ["姓名", "项目", "人员类别", "基本工资", "绩效工资"],
                [["甲", "A", "外聘", 5000, 1000], ["乙", "B", "外聘", 4000, 800]],
            )
            rows = calculate(
                [*read_rows(roster_a), *read_rows(roster_b)],
                read_rows(history),
                records, "2026-06", None, None,
            )
            self.assertEqual({row["姓名"] for row in rows["payroll_detail"]}, {"甲", "乙"})

    def test_payroll_rows_keep_each_project_contiguous(self):
        rows = calculate(
            [],
            [
                {"name": "乙", "project": "B项目", "category": "外包",
                 "position_salary": 2000, "performance": 1000},
                {"name": "甲", "project": "A项目", "category": "外包",
                 "position_salary": 2000, "performance": 1000},
                {"name": "丙", "project": "A项目", "category": "外包",
                 "position_salary": 2000, "performance": 1000},
            ],
            [
                {"name": "甲", "project": "A项目", "category": "外包", "actual_work_days": 21},
                {"name": "乙", "project": "B项目", "category": "外包", "actual_work_days": 21},
                {"name": "丙", "project": "A项目", "category": "外包", "actual_work_days": 21},
            ],
            "2026-06", None, None,
        )
        self.assertEqual([row["项目"] for row in rows["payroll_detail"]], ["A项目", "A项目", "B项目"])
        self.assertEqual([row["姓名"] for row in rows["payroll_detail"]], ["甲", "丙", "乙"])

    def test_conflicting_attendance_months_are_blocked(self):
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "june.json"
            second = Path(directory) / "july.json"
            first.write_text('{"month":"2026-06","records":[{"name":"甲","actual_work_days":21}]}', encoding="utf-8")
            second.write_text('{"month":"2026-07","records":[{"name":"乙","actual_work_days":21}]}', encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "考勤月份冲突"):
                attendance_records_from_paths([first, second])

    def test_formula_cells_without_excel_cache_supply_seniority_and_construction(self):
        with tempfile.TemporaryDirectory() as directory:
            history = self.make_book(
                directory, "B项目2026年5月外包人员工资表.xlsx",
                ["姓名", "项目", "人员类别", "工作天数", "基本工资", "绩效工资", "工龄工资", "施工补贴"],
                [["李雷", "B项目", "外包", 26, 3200, 4200, "=30*6+30+30", "=D2*30"]],
            )
            rows = calculate(
                [{"name": "李雷（大）", "project": "B项目", "category": "外包", "status": "正常"}],
                read_rows(history),
                [{"name": "李雷", "project": "B项目", "category": "外包", "actual_work_days": 26}],
                "2026-06", None, None,
            )
            detail = rows["payroll_detail"][0]
            self.assertEqual(detail["工龄工资"], 240)
            self.assertEqual(detail["施工补贴"], 780)

    def test_filename_is_not_used_for_project_or_category(self):
        path = Path("B项目2026年5月甲司外包人员工资表.xlsx")
        self.assertEqual(source_hints(path), ("", ""))
        with tempfile.TemporaryDirectory() as directory:
            history = self.make_book(
                directory, "B项目2026年5月甲司外包人员工资表.xlsx",
                ["姓名", "基本工资", "绩效工资"],
                [["张三", 5000, 1000]],
            )
            loaded = read_rows(history)
            self.assertFalse(loaded[0].get("project"))
            self.assertFalse(loaded[0].get("_project_hint"))

    def test_filename_vendor_short_name_matches_outsourced_label(self):
        with tempfile.TemporaryDirectory() as directory:
            history = self.make_book(
                directory, "A项目2026年5月甲司人员工资表.xlsx",
                ["姓名", "基本工资", "绩效工资"],
                [["张三", 5000, 1000]],
            )
            rows = calculate(
                [],
                read_rows(history),
                [{"name": "张三", "project": "A项目", "category": "甲司外包",
                  "actual_work_days": 21}],
                "2026-06", None, None,
            )
            self.assertEqual(rows["payroll_detail"][0]["岗位工资"], 5000)
            self.assertEqual(rows["payroll_detail"][0]["人员类别"], "")

    def test_filename_vendor_short_name_does_not_match_other_vendor(self):
        rows = calculate(
            [],
            [{"name": "张三", "project": "A项目", "category": "甲司",
              "position_salary": 5000, "performance": 1000}],
            [{"name": "张三", "project": "A项目", "category": "乙司外包",
              "actual_work_days": 21}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"], [])
        self.assertTrue(any("未匹配历史工资" in item["说明"] for item in rows["review_exceptions"]))

    def test_generic_filename_category_still_matches_specific_history_label(self):
        rows = calculate(
            [],
            [{"name": "张三", "project": "A项目", "category": "外包",
              "position_salary": 5000, "performance": 1000}],
            [{"name": "张三", "project": "A项目", "category": "甲司外包",
              "actual_work_days": 21}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["岗位工资"], 5000)
        self.assertEqual(rows["payroll_detail"][0]["人员类别"], "外包")

    def test_submitted_cross_month_history_is_used_for_reconciliation(self):
        rows = calculate(
            [],
            [{"name": "历史对比", "project": "B项目", "category": "外包",
              "position_salary": 5000, "performance": 1000, "historical_gross": 6000,
              "_source_month": "2026-05"}],
            [{"name": "历史对比", "project": "B项目", "category": "外包", "actual_work_days": 21}],
            "2026-06", None, None,
        )
        comparison = rows["reconciliation"][0]
        self.assertEqual(comparison["历史应发"], 6000)
        self.assertEqual(comparison["状态"], "一致")

    def test_official_holiday_is_retained_when_vision_only_returns_weekends(self):
        rows = calculate(
            [],
            [{"name": "端午", "project": "B项目", "category": "外包",
              "position_salary": 3800, "performance": 1000, "overtime_amount": 1400}],
            [{"name": "端午", "project": "B项目", "category": "外包",
              "actual_work_days": 25,
              "holiday_dates": ["2026-06-20", "2026-06-21"]}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["加班天数"], 4)

    def test_leave_days_is_not_personal_leave(self):
        history = [{"name": "请假", "project": "B项目", "category": "外包",
                    "position_salary": 5000, "performance": 2175}]
        rows = calculate(
            [], history,
            [{"name": "请假", "project": "B项目", "category": "外包",
              "actual_work_days": 21, "leave_days": 2, "comp_leave_days": 3}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["事假天数"], 0)
        self.assertEqual(rows["payroll_detail"][0]["绩效工资"], 2175)

    def test_sparse_excel_used_range_does_not_scan_the_whole_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "B项目2026年5月外包人员工资表.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["姓名", "项目", "人员类别", "基本工资", "绩效工资"])
            sheet.append(["张三", "B项目", "外包", 3200, 4200])
            sheet.cell(row=50000, column=1, value="远行")
            workbook.save(path)
            started = time.perf_counter()
            rows = read_rows(path)
            self.assertLess(time.perf_counter() - started, 2)
            names = [row.get("name") for row in rows]
            self.assertIn("张三", names)
            self.assertNotIn("远行", names)

    def test_personal_leave_requires_shi_mark_not_month_remainder(self):
        history = [{"name": "周七", "project": "C项目", "category": "外包",
                    "position_salary": 3200, "performance": 4200, "overtime_amount": 848}]
        remainder = calculate(
            [], history,
            [{"name": "周七", "project": "C项目", "category": "外包",
              "actual_work_days": 26, "leave_days": 4}],
            "2026-06", None, None,
        )
        self.assertEqual(remainder["payroll_detail"][0]["事假天数"], 0)
        self.assertEqual(remainder["payroll_detail"][0]["绩效工资"], 4200)
        self.assertEqual(remainder["payroll_detail"][0]["加班天数"], 4)

        rest_marks = calculate(
            [], history,
            [{"name": "周七", "project": "C项目", "category": "外包",
              "actual_work_days": 26,
              "personal_leave_days": 4,
              "daily_marks": ["8"] * 26 + ["休", "换", "休", "调"]}],
            "2026-06", None, None,
        )
        self.assertEqual(rest_marks["payroll_detail"][0]["事假天数"], 0)
        self.assertEqual(rest_marks["payroll_detail"][0]["绩效工资"], 4200)

        marked = calculate(
            [], history,
            [{"name": "周七", "project": "C项目", "category": "外包",
              "actual_work_days": 21,
              "personal_leave_days": 9,
              "daily_marks": ["8"] * 21 + ["事", "事"]}],
            "2026-06", None, None,
        )
        self.assertEqual(marked["payroll_detail"][0]["事假天数"], 2)
        self.assertEqual(marked["payroll_detail"][0]["绩效工资"], 4200 - 4200 / 21.75 * 2)

    def test_consecutive_personal_leave_is_kept_even_if_plus_attendance_fills_month(self):
        rows = calculate(
            [],
            [{"name": "王自鑫", "position_salary": 2100, "performance": 2700}],
            [{"name": "王自鑫", "actual_work_days": 17, "marked_work_days": 17,
              "personal_leave_days": 13, "note": "事假"}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["事假天数"], 13)
        self.assertAlmostEqual(
            rows["payroll_detail"][0]["绩效工资"],
            2700 - 2700 / 21.75 * 13,
        )

    def test_printed_attendance_must_equal_mark_count(self):
        history = [{"name": "对账", "project": "C项目", "category": "外包",
                    "position_salary": 2000, "performance": 1000, "overtime_amount": 184}]
        mismatched = calculate(
            [], history,
            [{"name": "对账", "project": "C项目", "category": "外包",
              "actual_work_days": 25, "marked_work_days": 26}],
            "2026-06", None, None,
        )
        self.assertEqual(mismatched["payroll_detail"][0]["实际出勤"], 25)
        self.assertTrue(any("印刷实际出勤与出勤符号天数不一致" in item["说明"]
                            for item in mismatched["review_exceptions"]))

        matched = calculate(
            [], history,
            [{"name": "对账", "project": "C项目", "category": "外包",
              "actual_work_days": 26, "marked_work_days": 26}],
            "2026-06", None, None,
        )
        self.assertFalse(any("印刷实际出勤与出勤符号天数不一致" in item["说明"]
                             for item in matched["review_exceptions"]))

        partial = calculate(
            [], history,
            [{"name": "对账", "project": "C项目", "category": "外包",
              "actual_work_days": 27,
              "daily_marks": {"2026-06-06": "出勤", "2026-06-07": "出勤"}}],
            "2026-06", None, None,
        )
        self.assertFalse(any("印刷实际出勤与出勤符号天数不一致" in item["说明"]
                             for item in partial["review_exceptions"]))

    def test_generated_result_workbook_is_rejected_as_sole_history(self):
        with tempfile.TemporaryDirectory() as directory:
            generated = Path(directory) / "2026年6月工资核算表.xlsx"
            write_workbook(
                calculate(
                    [],
                    [{"name": "张三", "project": "A项目", "category": "外聘",
                      "position_salary": 5000, "performance": 1000, "historical_gross": 6000}],
                    [{"name": "张三", "project": "A项目", "category": "外聘",
                      "actual_work_days": 21}],
                    "2026-06", None, None,
                ),
                generated,
            )
            with self.assertRaisesRegex(ValueError, "系统生成的工资核算表"):
                load_historical_payroll([generated])

    def test_generated_result_workbook_is_skipped_when_mixed_with_human_history(self):
        with tempfile.TemporaryDirectory() as directory:
            human = self.make_book(
                directory, "2026年5月人工工资表.xlsx",
                ["姓名", "项目", "人员类别", "基本工资", "绩效工资", "应发工资"],
                [["张三", "A项目", "外聘", 5000, 1000, 6800]],
            )
            generated = Path(directory) / "2026年6月工资核算表.xlsx"
            write_workbook(
                calculate(
                    [],
                    [{"name": "张三", "project": "A项目", "category": "外聘",
                      "position_salary": 9999, "performance": 1, "historical_gross": 1}],
                    [{"name": "张三", "project": "A项目", "category": "外聘",
                      "actual_work_days": 21}],
                    "2026-06", None, None,
                ),
                generated,
            )
            rows, skipped = load_historical_payroll([generated, human])
            self.assertEqual(skipped, ["2026年6月工资核算表.xlsx"])
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["position_salary"], 5000)
            result = calculate(
                [],
                rows,
                [{"name": "张三", "project": "A项目", "category": "外聘",
                  "actual_work_days": 21}],
                "2026-06", None, None,
            )
            self.assertEqual(result["reconciliation"][0]["历史应发"], 6800)
            self.assertNotEqual(result["reconciliation"][0]["状态"], "历史应发不唯一")
            notes = note_skipped_generated_history(result, skipped)
            self.assertEqual(notes["skipped_generated_historical"], skipped)
            self.assertTrue(any("错传或漏传" in item["说明"] for item in result["review_exceptions"]))
            attendance = Path(directory) / "attendance.json"
            attendance.write_text(
                '{"month":"2026-06","records":[{"name":"张三","project":"A项目","category":"外聘","actual_work_days":21}]}',
                encoding="utf-8",
            )
            roster = self.make_book(
                directory, "roster.xlsx", ["姓名", "项目", "人员类别"],
                [["张三", "A项目", "外聘"]],
            )
            output = Path(directory) / "out.xlsx"
            buffer = io.StringIO()
            with redirect_stdout(buffer):
                code = main([
                    "--roster", str(roster),
                    "--historical", str(generated), str(human),
                    "--attendance", str(attendance),
                    "--month", "2026-06",
                    "--output", str(output),
                ])
            self.assertEqual(code, 0)
            payload = json.loads(buffer.getvalue())
            self.assertEqual(payload["skipped_generated_historical"], ["2026年6月工资核算表.xlsx"])
            self.assertTrue(any("错传或漏传" in warning for warning in payload["warnings"]))
            self.assertEqual(payload["statutory_holidays"], ["2026-06-19", "2026-06-20", "2026-06-21"])

    def test_parse_construction_formula_terms(self):
        self.assertEqual(parse_construction_formula("=D11*30"), [(None, 30.0)])
        self.assertEqual(parse_construction_formula("=15*30+11*50"), [(15.0, 30.0), (11.0, 50.0)])

    def test_performance_formula_uses_standard_not_prorated_amount(self):
        self.assertEqual(parse_performance_standard("=ROUND(2700/21.75*17,2)"), 2700)
        self.assertEqual(parse_performance_standard("=2700-2700/21.75*5"), 2700)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["姓名", "工作天数", "基本工资", "绩效工资"])
            sheet.append(["王自鑫", 17, 2100, "=ROUND(2700/21.75*17,2)"])
            workbook.save(path)
            loaded = read_rows(path)
            self.assertEqual(loaded[0]["performance"], 2700)
            rows = calculate(
                [], loaded,
                [{"name": "王自鑫", "actual_work_days": 17, "personal_leave_days": 5,
                  "daily_marks": {str(day): "事" for day in range(1, 6)} | {str(day): "8" for day in range(6, 23)}}],
                "2026-06", None, None,
            )
            self.assertEqual(rows["baseline"][0]["绩效工资"], 2700)
            self.assertEqual(rows["payroll_detail"][0]["事假天数"], 5)
            self.assertAlmostEqual(
                rows["payroll_detail"][0]["绩效工资"],
                2700 - 2700 / 21.75 * 5,
            )

    def test_segmented_construction_subsidy_from_formula_and_attendance(self):
        rows = calculate(
            [],
            [{"name": "刘志才", "category": "人事代理", "position_salary": 3000,
              "performance": 1000, "work_days": 26, "_construction_formula": "=15*30+11*50"}],
            [{"name": "刘志才", "category": "人事代理", "actual_work_days": 26,
              "project_segments": [
                  {"project": "杨职院", "actual_work_days": 15},
                  {"project": "榆林西", "actual_work_days": 11},
              ]}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["施工补贴"], 15 * 30 + 11 * 50)
        self.assertEqual(rows["payroll_detail"][0]["实际出勤"], 26)

    def test_yellow_trailing_subtotal_assigns_project_from_sheet_content(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["姓名", "岗位名称", "工作天数", "基本工资", "绩效工资", "施工补贴"])
            sheet.append(["张志宏", "施工员", 21, 2400, 2900, "=C2*20"])
            yellow = PatternFill("solid", fgColor="FFFF00")
            sheet.append(["4-R2", "小计", "=SUM(C2:C2)", "=SUM(D2:D2)", "=SUM(E2:E2)", "=SUM(F2:F2)"])
            for cell in sheet[3]:
                cell.fill = yellow
            workbook.save(path)
            loaded = read_rows(path)
            self.assertEqual(loaded[0]["project"], "4-R2")
            rows = calculate(
                [], loaded,
                [{"name": "张志宏", "actual_work_days": 21}],
                "2026-06", None, None,
            )
            self.assertEqual(rows["payroll_detail"][0]["项目"], "4-R2")
            self.assertEqual(rows["payroll_detail"][0]["施工补贴"], 420)

    def test_blank_project_still_calculates_when_history_matches(self):
        rows = calculate(
            [],
            [{"name": "王五", "category": "外聘", "position_salary": 5000, "performance": 1000}],
            [{"name": "王五", "category": "外聘", "actual_work_days": 21}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["岗位工资"], 5000)
        self.assertEqual(rows["payroll_detail"][0]["项目"], "")

    def test_page_department_containing_job_title_stays_a_project(self):
        yangling = "杨凌职业技术学院新校区项目经理部"
        rows = calculate(
            [],
            [{"name": "付浩", "project": "宇航钛合金智能锻造项目", "category": "外聘",
              "position_salary": 3000, "performance": 1000}],
            [{"name": "付浩", "category": "外聘", "actual_work_days": 26,
              "page_project": yangling}],
            "2026-06", None, None,
            attendance_project="宇航级钛及钛合金智能锻造产线及供应链协同建设项目",
        )
        self.assertEqual(rows["payroll_detail"][0]["项目"], yangling)

    def test_mixed_attendance_files_do_not_share_first_page_project(self):
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "page-1.json"
            second = Path(directory) / "page-10.json"
            first.write_text(json.dumps({
                "month": "2026-06",
                "project": "宇航级钛及钛合金智能锻造产线及供应链协同建设项目",
                "records": [{"name": "樊云刚", "actual_work_days": 21}],
            }), encoding="utf-8")
            second.write_text(json.dumps({
                "month": "2026-06",
                "project": "杨凌职业技术学院新校区项目经理部",
                "records": [{"name": "付浩", "actual_work_days": 26}],
            }), encoding="utf-8")
            _month, shared, records = attendance_records_from_paths([first, second])
            self.assertIsNone(shared)
            rows = calculate(
                [],
                [{"name": "樊云刚", "position_salary": 3000, "performance": 1000},
                 {"name": "付浩", "position_salary": 3000, "performance": 1000}],
                records, "2026-06", None, None, None, shared,
            )
            details = {row["姓名"]: row["项目"] for row in rows["payroll_detail"]}
            self.assertIn("宇航", details["樊云刚"])
            self.assertIn("杨凌", details["付浩"])

    def test_yitong_and_yitong_homophone_are_distinct_vendors(self):
        rows = calculate(
            [],
            [{"name": "张三", "category": "易通外包", "position_salary": 5000, "performance": 1000},
             {"name": "张三", "category": "益通外包", "position_salary": 4000, "performance": 800}],
            [{"name": "张三", "category": "易通外包", "actual_work_days": 21},
             {"name": "张三", "category": "益通外包", "actual_work_days": 22}],
            "2026-06", None, None,
        )
        details = {(row["人员类别"], row["岗位工资"]) for row in rows["payroll_detail"]}
        self.assertEqual(details, {("易通外包", 5000), ("益通外包", 4000)})

    def test_known_rest_and_maternity_marks_are_not_unrecognized(self):
        rows = calculate(
            [],
            [{"name": "李珍", "category": "博途外包", "position_salary": 2000, "performance": 2500,
              "work_days": 0, "_construction_formula": "=(D2*30)*0"}],
            [{"name": "李珍", "category": "博途外包", "actual_work_days": 0,
              "unrecognized_marks": ["产", "走", "加班"],
              "daily_marks": ["产"] * 30}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["施工补贴"], 0)
        self.assertFalse(any("未识别考勤符号" in item["说明"] for item in rows["review_exceptions"]))
        self.assertTrue(any("产" in item["说明"] for item in rows["review_exceptions"]))

    def test_colleague_rate_fills_cross_project_without_own_formula(self):
        rows = calculate(
            [],
            [
                {"name": "同事", "project": "A项目", "category": "外包",
                 "position_salary": 2000, "performance": 1000, "_construction_formula": "=D2*30"},
                {"name": "同事乙", "project": "B项目", "category": "外包",
                 "position_salary": 2000, "performance": 1000, "_construction_formula": "=D3*50"},
                {"name": "本人", "project": "A项目", "category": "外包",
                 "position_salary": 2000, "performance": 1000},
            ],
            [{"name": "本人", "category": "外包", "actual_work_days": 26,
              "project_segments": [
                  {"project": "A项目", "actual_work_days": 10},
                  {"project": "B项目", "actual_work_days": 16},
              ]}],
            "2026-06", None, None,
        )
        self.assertEqual(rows["payroll_detail"][0]["施工补贴"], 10 * 30 + 16 * 50)

    def test_output_workbook_writes_calculation_formulas(self):
        rows = calculate(
            [],
            [{"name": "赵六", "project": "C", "category": "外聘",
              "work_days": 26, "position_salary": 5000, "performance": 2175,
              "construction_day": 30}],
            [{"name": "赵六", "project": "C", "category": "外聘",
              "actual_work_days": 21, "daily_marks": ["8"] * 21 + ["事"]}],
            "2026-06", None, None,
        )
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            write_workbook(rows, output)
            workbook = load_workbook(output, data_only=False)
            try:
                sheet = workbook["工资核算明细"]
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                values = {headers[i]: sheet.cell(2, i + 1).value for i in range(len(headers))}
                self.assertTrue(str(values["绩效工资"]).startswith("=MAX("))
                self.assertTrue(str(values["施工补贴"]).startswith("="))
                self.assertTrue(str(values["应发工资"]).startswith("="))
                self.assertIn("F2", str(values["应发工资"]))
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
