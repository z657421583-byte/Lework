#!/usr/bin/env python3
"""外聘人员确定性工资计算器。

仅读取本地 xlsx/JSON，不联网、不读取环境变量中的密钥，也不修改输入文件。
"""

from __future__ import annotations

import argparse
import ast
import calendar
import difflib
import json
import math
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.cell import column_index_from_string
except ImportError as exc:  # pragma: no cover - exercised by CLI environments
    raise SystemExit("需要安装 openpyxl：python -m pip install openpyxl") from exc

ERROR_INPUT = 2
ERROR_BLOCKED = 3
ERROR_OUTPUT = 4
MONEY = "0.00"
# 人工工资表格式化整列时，openpyxl 会把 used range 扩到十几万行。
MAX_SHEET_ROWS = 2000
MAX_SHEET_COLS = 48
EMPTY_ROW_STOP = 40

# Workbook sheet titles written by this calculator. A file that contains all of
# them is a generated result, not a human historical payroll sheet.
GENERATED_PAYROLL_SHEETS = (
    "工资核算明细",
    "工资基准",
    "考勤汇总",
    "历史工资对比",
    "人工复核事项",
)

# 国务院办公厅发布的 2026 年放假调休安排。考勤识别没有返回节假日时使用；
# 新年份应在公告发布后补充，避免把非工作日猜成工作日。
OFFICIAL_CALENDAR = {
    2026: {
        "holidays": {
            "2026-01-01", "2026-01-02", "2026-01-03",
            "2026-02-15", "2026-02-16", "2026-02-17", "2026-02-18", "2026-02-19",
            "2026-02-20", "2026-02-21", "2026-02-22", "2026-02-23",
            "2026-04-04", "2026-04-05", "2026-04-06",
            "2026-05-01", "2026-05-02", "2026-05-03", "2026-05-04", "2026-05-05",
            "2026-06-19", "2026-06-20", "2026-06-21",
            "2026-09-25", "2026-09-26", "2026-09-27",
            "2026-10-01", "2026-10-02", "2026-10-03", "2026-10-04", "2026-10-05",
            "2026-10-06", "2026-10-07",
        },
        "working_weekends": {
            "2026-01-04", "2026-02-14", "2026-02-28", "2026-05-09",
            "2026-09-20", "2026-10-10",
        },
    },
}

YELLOW_FILLS = (
    "FFFF00", "FFFF99", "FFC000", "FFEB9C", "FFF2CC", "FFFFCC",
    "FFD966", "FFEA00", "F4B183", "FFFF66",
)

KNOWN_ATTENDANCE_MARKS = {
    "8", "出勤", "上班", "正常", "加班", "周末加班", "节假日加班",
    "√", "✔", "1", "休", "换", "调", "走", "年假", "/", "事", "事假",
    "病", "旷", "产", "产假", "陪产", "陪产假", "探", "探亲", "探亲假",
    "婚", "婚假", "公", "公假", "学", "学习",
}

REVIEW_LEAVE_MARKS = {
    "产", "产假", "陪产", "陪产假", "探", "探亲", "探亲假",
    "婚", "婚假", "公", "公假", "学", "学习",
}

ALIASES = {
    "name": ("姓名", "员工姓名", "人员姓名", "名字"),
    "project": ("项目", "部门", "部门/项目", "部门项目", "项目名称", "所属项目"),
    "category": ("人员类别", "类别", "人员类型", "用工类别"),
    "status": ("状态", "在册状态", "人员状态"),
    "position": ("岗位", "岗位名称", "职务"),
    "position_salary": ("基本工资", "岗位工资", "工资标准"),
    "performance": ("绩效工资", "绩效", "绩效标准"),
    "seniority": ("工龄工资", "工龄"),
    "title": ("职称工资", "职称"),
    "construction_day": ("施工补贴日标准", "施工补贴标准", "施工补贴/天"),
    "construction": ("施工补贴", "施工补贴日标准"),
    "transport": ("交通补助", "交通补贴"),
    "phone": ("话费", "话费补贴"),
    "phone_1_3": ("1-3月话补", "1－3月话补"),
    "phone_4_6": ("4-6月话补", "4－6月话补"),
    "phone_7_9": ("7-9月话补", "7－9月话补"),
    "phone_10_12": ("10-12月话补", "10－12月话补"),
    "hot": ("降温费", "高温补贴"),
    "overtime_standard": ("双休日加班标准", "加班标准", "双休日加班（标准）"),
    "overtime_count": ("双休日加班个数", "加班个数", "双休日加班（个数）"),
    "overtime_amount": ("双休日加班金额", "加班金额", "双休日加班（金额）", "双休日加班"),
    "historical_gross": ("应发工资", "人工应发工资", "应发"),
    "work_days": ("工作天数", "实际出勤", "出勤天数"),
}

OVERTIME_SUBHEADERS = ("标准", "个数", "金额")


def clean(value: Any) -> str:
    text = unicodedata.normalize("NFKC", "" if value is None else str(value))
    return re.sub(r"[\s\u3000]+", "", text).strip()


def normalized_name(value: Any) -> str:
    """Normalize a person name without treating trailing roster annotations as identity."""
    return re.sub(r"[（(][^（）()]*[）)]$", "", clean(value)).strip()


def name_annotation(value: Any) -> str:
    """Return a trailing roster annotation for transparent fuzzy-match evidence."""
    match = re.search(r"[（(]([^（）()]*)[）)]$", clean(value))
    return match.group(1) if match else ""


def number(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def key(name: Any, project: Any, category: Any) -> tuple[str, str, str]:
    return normalized_name(name), clean(project), clean(category)


def is_unknown_name(value: Any) -> bool:
    return clean(value) in {"未知", "未知人员", "无法识别", "不详", "无名"}


JOB_TITLES = {
    "项目经理", "项目副经理", "技术负责人", "施工员", "技术员", "质量员",
    "资料员", "安全员", "预算员", "商务经理", "后勤", "试验", "电工",
    "司机", "厨师", "保洁", "生产经理", "现场经理", "装修主管", "安全主管",
}


def is_job_title(value: Any) -> bool:
    """岗位名须整段匹配。项目名里的「项目经理部」不能当成岗位。"""
    text = clean(value)
    if not text:
        return False
    if text in JOB_TITLES:
        return True
    return any(text.endswith(title) and len(text) <= len(title) + 4 for title in JOB_TITLES)


def is_employment_category(value: str) -> bool:
    """Return whether a label can safely constrain an external-worker match."""
    return any(token in value for token in ("外包", "派遣", "劳务", "外聘", "人事代理"))


def looks_like_department_or_title(value: str) -> bool:
    """Vision often puts a department or job title into the category field."""
    return is_job_title(value) or "管理" in value


def is_summary(row: dict[str, Any]) -> bool:
    text = "".join(clean(v) for v in row.values() if v is not None)
    return not text or any(word in text for word in ("小计", "合计", "总计", "汇总"))


def looks_like_overtime_subheader(values: list[Any] | tuple[Any, ...]) -> bool:
    """第二行只有「标准/个数/金额」时，与上一行「双休日加班」拼成完整表头。"""
    labels = [clean(value) for value in values if clean(value)]
    return bool(labels) and all(label in OVERTIME_SUBHEADERS for label in labels) and len(labels) >= 2


def compose_two_row_headers(top: list[Any], sub: list[Any] | tuple[Any, ...]) -> list[str]:
    """合并单元格只留左上角文字：向下填充父表头，再拼子表头。"""
    filled: list[str] = []
    last = ""
    width = max(len(top), len(sub))
    for index in range(width):
        text = clean(top[index]) if index < len(top) else ""
        if text:
            last = text
        filled.append(last)
    composed: list[str] = []
    for index, parent in enumerate(filled):
        child = clean(sub[index]) if index < len(sub) else ""
        if child in OVERTIME_SUBHEADERS:
            composed.append(f"{parent}{child}" if parent else child)
        elif child and not parent:
            composed.append(child)
        else:
            composed.append(parent)
    return composed


def header_map(headers: list[Any]) -> dict[str, int]:
    normalized = [clean(h) for h in headers]
    result = {}
    for field, names in ALIASES.items():
        for alias in names:
            if clean(alias) in normalized:
                result[field] = normalized.index(clean(alias))
                break
    for field, pattern in (
        ("phone_1_3", r"1(?:月)?[-－—至]?3月话补"),
        ("phone_4_6", r"4(?:月)?[-－—至]?6月话补"),
        ("phone_7_9", r"7(?:月)?[-－—至]?9月话补"),
        ("phone_10_12", r"10(?:月)?[-－—至]?12月话补"),
    ):
        if field not in result:
            for index, header in enumerate(normalized):
                if re.search(pattern, header):
                    result[field] = index
                    break
    return result


def cell_fill_rgb(cell: Any) -> str:
    fill = getattr(cell, "fill", None)
    if not fill or fill.fill_type in (None, "none"):
        return ""
    color = fill.fgColor or fill.start_color
    rgb = str(getattr(color, "rgb", "") or "")
    return rgb.upper()


def cell_is_yellow(cell: Any) -> bool:
    rgb = cell_fill_rgb(cell)
    return any(token in rgb for token in YELLOW_FILLS)


def looks_like_project_label(value: Any) -> bool:
    text = clean(value)
    if not text or any(word in text for word in ("合计", "总计", "小计", "序号", "姓名")):
        return False
    if is_job_title(text) or is_employment_category(text):
        return False
    return True


def parse_performance_standard(formula: Any) -> float | None:
    """绩效单元格常是 ROUND(标准/21.75*出勤,2) 或 标准-标准/21.75*事假。返回标准，不用折算后的值。"""
    text = str(formula or "").strip()
    if not text.startswith("=") or "21.75" not in text.replace(" ", ""):
        return None
    expr = re.sub(r"\s+", "", text[1:]).upper()
    numbers = [float(token) for token in re.findall(r"\d+(?:\.\d+)?", expr)]
    standards = [
        value for value in numbers
        if value > 31 and not math.isclose(value, 21.75, abs_tol=1e-6)
    ]
    unique = {round(value, 4) for value in standards}
    if len(unique) == 1:
        return next(iter(unique))
    return None


def parse_construction_formula(formula: Any) -> list[tuple[float | None, float]]:
    """Parse workdays×rate terms from a construction-subsidy Excel formula."""
    text = str(formula or "").strip()
    if not text.startswith("="):
        return []
    expr = re.sub(r"\s+", "", text[1:]).upper()
    if "ROUND(" in expr or "SUM(" in expr:
        return []
    terms: list[tuple[float | None, float]] = []
    for days_token, rate_token in re.findall(
        r"([A-Z]{1,3}\d+|\d+(?:\.\d+)?)\*(\d+(?:\.\d+)?)",
        expr,
    ):
        rate = number(rate_token)
        if rate is None:
            continue
        if re.fullmatch(r"[A-Z]{1,3}\d+", days_token):
            terms.append((None, rate))
        else:
            terms.append((float(days_token), rate))
    return terms


def construction_formula_zeroed(formula: Any) -> bool:
    text = re.sub(r"\s+", "", str(formula or "").upper())
    return bool(re.search(r"\)\s*\*\s*0(?:\.0+)?$", text)) or text.endswith("*0")


def source_hints(_path: Path) -> tuple[str, str]:
    """Filename is not a stable source of project or employment category."""
    return "", ""


def source_month(path: Path, sheet_name: str = "") -> str | None:
    """Extract a month from a source name without inferring a payroll month."""
    text = f"{path.name} {sheet_name}"
    match = re.search(r"(20\d{2})\s*[年\-/\.]\s*(\d{1,2})\s*月?", text)
    if not match:
        return None
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def evaluate_excel_formula(
    sheet: Any,
    expression: str,
    cache: dict[str, float | None],
) -> float | None:
    """Evaluate the simple arithmetic formulas used by historical payroll sheets."""
    cell_ref_pattern = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?(\d+)")

    def cell_value(match: re.Match[str]) -> str:
        coordinate = f"{match.group(1)}{match.group(2)}"
        if coordinate not in cache:
            cache[coordinate] = None
            cache[coordinate] = formula_cell_number(
                sheet,
                int(match.group(2)),
                column_index_from_string(match.group(1)),
                cache,
            )
        value = cache[coordinate]
        return str(value) if value is not None else "0"

    normalized = cell_ref_pattern.sub(cell_value, expression[1:].upper())
    normalized = normalized.replace("^", "**").replace("ROUND(", "round(")
    try:
        tree = ast.parse(normalized, mode="eval")
    except SyntaxError:
        return None
    allowed_nodes = (
        ast.Expression, ast.BinOp, ast.UnaryOp, ast.Add, ast.Sub, ast.Mult,
        ast.Div, ast.Pow, ast.USub, ast.UAdd, ast.Constant, ast.Call, ast.Name,
        ast.Load,
    )
    if any(not isinstance(node, allowed_nodes) for node in ast.walk(tree)):
        return None
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and node.id not in {"round", "min", "max"}:
            return None
        if isinstance(node, ast.Call) and not isinstance(node.func, ast.Name):
            return None
    try:
        value = eval(compile(tree, "<payroll-formula>", "eval"), {"__builtins__": {}}, {
            "round": round, "min": min, "max": max,
        })
    except (ArithmeticError, TypeError, ValueError):
        return None
    return number(value)


def formula_cell_number(
    sheet: Any,
    row: int,
    column: int,
    cache: dict[str, float | None],
) -> float | None:
    """Resolve a numeric cell, including a same-sheet arithmetic formula."""
    cell = sheet.cell(row=row, column=column)
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        return evaluate_excel_formula(sheet, value, cache)
    return number(value)


def is_blank_row(values: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def worksheet_values(sheet: Any) -> list[tuple[Any, ...]]:
    """Return display values while preserving formulas without cached Excel results."""
    cache: dict[str, float | None] = {}
    rows: list[tuple[Any, ...]] = []
    empty_streak = 0
    max_row = min(sheet.max_row or 1, MAX_SHEET_ROWS)
    max_col = min(sheet.max_column or 1, MAX_SHEET_COLS)
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        values = []
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                resolved = formula_cell_number(sheet, cell.row, cell.column, cache)
                values.append(resolved if resolved is not None else value)
            else:
                values.append(value)
        if is_blank_row(values):
            empty_streak += 1
            if empty_streak >= EMPTY_ROW_STOP and any(not is_blank_row(item) for item in rows):
                break
            rows.append(tuple(values))
            continue
        empty_streak = 0
        rows.append(tuple(values))
    return rows


def is_generated_payroll_workbook(path: Path) -> bool:
    """Return whether the workbook is this calculator's multi-sheet result."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        names = set(workbook.sheetnames)
    finally:
        workbook.close()
    return set(GENERATED_PAYROLL_SHEETS).issubset(names)


def load_historical_payroll(paths: list[Path]) -> tuple[list[dict[str, Any]], list[str]]:
    """Load human historical payroll rows and skip generated result workbooks.

    Generated 工资核算表 files repeat the same people across five sheets, which
    would make every baseline match ambiguous. They are not a substitute for
    human historical payroll. If every supplied file is generated, fail instead
    of producing an empty comparison.
    """
    generated: list[str] = []
    usable: list[Path] = []
    for path in paths:
        if is_generated_payroll_workbook(path):
            generated.append(path.name)
        else:
            usable.append(path)
    if not usable:
        names = "、".join(generated) or "未指定文件"
        raise ValueError(
            f"历史工资表「{names}」是系统生成的工资核算表，不能作为历史人工工资表。"
            "请改传往期人工编制的工资表。"
        )
    rows = [row for path in usable for row in read_rows(path)]
    return rows, generated


def note_skipped_generated_history(
    result: dict[str, list[dict[str, Any]]],
    skipped: list[str],
) -> dict[str, Any]:
    """Record skipped generated workbooks for the Excel review sheet and CLI JSON."""
    if not skipped:
        return {"skipped_generated_historical": [], "warnings": []}
    listed = "、".join(f"「{name}」" for name in skipped)
    warning = (
        f"已忽略系统生成的工资核算表{listed}，未作为历史人工工资表。"
        "本次工资基准和历史对比只使用其余人工历史工资表；请核对是否错传或漏传。"
    )
    result["review_exceptions"].append({
        "姓名": "", "项目": "", "人员类别": "", "类型": "输入资料",
        "说明": warning,
    })
    return {"skipped_generated_historical": skipped, "warnings": [warning]}


def projects_compatible(left: str, right: str) -> bool:
    if not left or not right:
        return True
    return left in right or right in left


def assign_yellow_block_projects(sheet: Any, people: list[dict[str, Any]]) -> None:
    """Attach project names from yellow header/subtotal rows to person rows."""
    if not people:
        return
    markers: list[tuple[int, str, str]] = []
    max_row = min(sheet.max_row or 1, MAX_SHEET_ROWS)
    max_col = min(sheet.max_column or 1, MAX_SHEET_COLS)
    for excel_row in range(1, max_row + 1):
        first = sheet.cell(excel_row, 1)
        label = first.value
        row_text = "".join(
            clean(sheet.cell(excel_row, col).value)
            for col in range(1, max_col + 1)
            if sheet.cell(excel_row, col).value is not None
        )
        if any(word in row_text for word in ("合计", "总计")) and "小计" not in clean(label):
            continue
        yellow = cell_is_yellow(first) or any(
            cell_is_yellow(sheet.cell(excel_row, col)) for col in range(1, min(4, max_col + 1))
        )
        has_subtotal = "小计" in row_text
        has_sum_above = False
        for col in range(1, max_col + 1):
            value = sheet.cell(excel_row, col).value
            if isinstance(value, str) and value.startswith("=") and "SUM(" in value.upper():
                numbers = [int(item) for item in re.findall(r"[A-Z]{1,3}(\d+)", value.upper())]
                if numbers and max(numbers) < excel_row:
                    has_sum_above = True
        if not looks_like_project_label(label):
            continue
        if yellow or has_subtotal:
            kind = "trailing" if has_subtotal or has_sum_above else "leading"
            markers.append((excel_row, clean(label), kind))
    if not markers:
        return
    for person in people:
        row_number = int(person.get("_row") or 0)
        project = ""
        previous: tuple[int, str, str] | None = None
        for marker in markers:
            if marker[0] > row_number:
                if marker[2] == "trailing":
                    project = marker[1]
                elif previous and previous[2] == "leading":
                    project = previous[1]
                break
            previous = marker
        else:
            if previous and previous[2] == "leading" and previous[0] < row_number:
                project = previous[1]
        if project:
            person["_block_project"] = project
            if not clean(person.get("project")):
                person["project"] = project


def read_rows(path: Path) -> list[dict[str, Any]]:
    """读取所有工作表，允许表头前有标题行。"""
    workbook = load_workbook(path, data_only=False, read_only=False)
    rows: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            raw = worksheet_values(sheet)
            header_index = next(
                (i for i, row in enumerate(raw) if len(header_map(list(row))) >= 2), None
            )
            if header_index is None:
                continue
            headers = list(raw[header_index])
            data_index = header_index + 1
            if data_index < len(raw) and looks_like_overtime_subheader(raw[data_index]):
                headers = compose_two_row_headers(headers, raw[data_index])
                data_index += 1
            mapping = header_map(headers)
            sheet_rows: list[dict[str, Any]] = []
            for row_number, values in enumerate(raw[data_index:], data_index + 1):
                row = {field: values[index] if index < len(values) else None
                       for field, index in mapping.items()}
                row["_sheet"] = sheet.title
                row["_row"] = row_number
                row["_source_file"] = path.name
                row["_source_month"] = source_month(path, sheet.title)
                construction_index = mapping.get("construction")
                if construction_index is not None:
                    original = sheet.cell(row_number, construction_index + 1).value
                    if isinstance(original, str) and original.startswith("="):
                        row["_construction_formula"] = original
                performance_index = mapping.get("performance")
                if performance_index is not None:
                    original = sheet.cell(row_number, performance_index + 1).value
                    if isinstance(original, str) and original.startswith("="):
                        row["_performance_formula"] = original
                        standard = parse_performance_standard(original)
                        if standard is not None:
                            row["performance"] = standard
                if not is_summary(row) and not is_unknown_name(row.get("name")):
                    sheet_rows.append(row)
            assign_yellow_block_projects(sheet, sheet_rows)
            rows.extend(sheet_rows)
    finally:
        workbook.close()
    return rows


def content_project(row: dict[str, Any]) -> str:
    project = row.get("project") or row.get("_block_project") or row.get("page_project")
    if is_job_title(project) or clean(project) in {"部门/分公司", "一分司", "二分司", "三分司", "直属项目部"}:
        return ""
    return clean(project)


def row_identity(row: dict[str, Any]) -> tuple[str, str, str]:
    name = row.get("name")
    project = content_project(row)
    if is_unknown_name(name):
        return "", "", ""
    category = row.get("category")
    if looks_like_department_or_title(clean(category)):
        category = ""
    elif not is_employment_category(clean(category)):
        # Keep explicit vendor labels from roster/history content only.
        category = category if clean(category) else ""
    return key(name, project, category)


def employment_categories_compatible(left: str, right: str) -> bool:
    """Treat nested labels such as 外包/某司 vs 某司外包 as the same employment type."""
    if not left or not right:
        return True
    if looks_like_department_or_title(left) or looks_like_department_or_title(right):
        return True
    if left == right or left in right or right in left:
        return True
    return not (is_employment_category(left) or is_employment_category(right))


def identity_matches(observed: tuple[str, str, str], candidate: tuple[str, str, str]) -> bool:
    """Match by name and employment category; project is display-only and may be empty."""
    observed_name, candidate_name = observed[0], candidate[0]
    if not observed_name or observed_name != candidate_name:
        return False
    return employment_categories_compatible(observed[2], candidate[2])


def compatible_identity(observed: tuple[str, str, str], candidate: tuple[str, str, str]) -> bool:
    """Check employment-category constraints before fuzzy name matching."""
    return employment_categories_compatible(observed[2], candidate[2])


def fuzzy_name_score(observed: Any, candidate: Any) -> tuple[float, str]:
    """Return conservative name similarity and a human-readable matching reason."""
    left, right = normalized_name(observed), normalized_name(candidate)
    if not left or not right:
        return 0, ""
    if left == right:
        return 1.0, "姓名去除末尾括号备注后一致"
    ratio = difflib.SequenceMatcher(a=left, b=right).ratio()
    # OCR has a small, well-known set of visually similar Chinese characters.
    confusion_pairs = {frozenset(pair) for pair in (("利", "丽"), ("翌", "罡"))}
    if len(left) == len(right):
        differences = [
            frozenset((a, b)) for a, b in zip(left, right) if a != b
        ]
        if len(differences) == 1 and differences[0] in confusion_pairs:
            return 0.95, "单字 OCR 易混"
        if len(left) >= 3 and len(differences) == 1:
            return 0.85, "单字姓名近似"
    if ratio >= 0.8:
        return ratio, f"编辑相似度 {ratio:.0%}"
    return ratio, ""


def fuzzy_match_review_note(reason: str) -> str:
    text = (reason or "").strip()
    if text.startswith("姓名"):
        return f"{text}，已模糊匹配，需人工复核是否同一人"
    return f"姓名{text}，已模糊匹配，需人工复核是否同一人"


def fuzzy_candidates(
    observed: dict[str, Any],
    candidates: list[dict[str, Any]],
    required_identity: tuple[str, str, str] | None = None,
) -> list[tuple[float, str, dict[str, Any]]]:
    """Rank only project/category-compatible unresolved people; never guess ties."""
    observed_identity = required_identity or row_identity(observed)
    result = []
    for candidate in candidates:
        candidate_identity = row_identity(candidate)
        if not compatible_identity(observed_identity, candidate_identity):
            continue
        score, reason = fuzzy_name_score(observed.get("name"), candidate.get("name"))
        if score >= 0.8 and reason:
            result.append((score, reason, candidate))
    return sorted(result, key=lambda item: item[0], reverse=True)


def uniquely_fuzzy_matched(
    observed: dict[str, Any],
    candidates: list[dict[str, Any]],
    required_identity: tuple[str, str, str] | None = None,
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    """唯一模糊候选仍采用并复核；只有多个同分/接近候选才算歧义并跳过。"""
    ranked = fuzzy_candidates(observed, candidates, required_identity)
    if not ranked:
        return None, None, None
    score, reason, winner = ranked[0]
    if len(ranked) == 1:
        return winner, reason, None
    next_score = ranked[1][0]
    if score >= 0.9 and score - next_score >= 0.08:
        return winner, reason, None
    choices = "、".join(
        f"{row.get('name')}[{row.get('_source_file', row.get('_sheet', '未知来源'))}!{row.get('_row', '')}]"
        for _, _, row in ranked[:3]
    )
    return None, None, f"模糊匹配存在歧义，候选：{choices}"


def attendance_records(path: Path) -> tuple[str | None, str | None, list[dict[str, Any]]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return None, None, payload
    if not isinstance(payload, dict) or not isinstance(payload.get("records"), list):
        raise ValueError("attendance JSON 必须是 records 数组或记录数组")
    month = payload.get("month")
    project = payload.get("project")
    holiday_dates = payload.get("holiday_dates", [])
    records = payload["records"]
    page_project = clean(project) or None
    if isinstance(holiday_dates, list):
        records = [
            {**record, "holiday_dates": record.get("holiday_dates", holiday_dates),
             "page_project": record.get("page_project") or record.get("project") or page_project}
            if isinstance(record, dict) else record
            for record in records
        ]
    elif isinstance(records, list):
        records = [
            {**record, "page_project": record.get("page_project") or record.get("project") or page_project}
            if isinstance(record, dict) else record
            for record in records
        ]
    return str(month) if month else None, page_project, records


def attendance_records_from_paths(
    paths: list[Path],
) -> tuple[str | None, str | None, list[dict[str, Any]]]:
    """Load one or more attendance JSON files and reject conflicting months."""
    months: set[str] = set()
    projects: set[str] = set()
    records: list[dict[str, Any]] = []
    for path in paths:
        month, source_project, source_records = attendance_records(path)
        if month:
            months.add(month)
        if source_project:
            projects.add(source_project)
        records.extend(source_records)
    if len(months) > 1:
        raise ValueError("考勤月份冲突：" + "、".join(sorted(months)))
    shared_project = next(iter(projects), None) if len(projects) == 1 else None
    return next(iter(months), None), shared_project, records


def attendance_mark_values(record: dict[str, Any]) -> list[str]:
    marks = record.get("daily_marks", record.get("days", []))
    if isinstance(marks, dict):
        marks = list(marks.values())
    if not isinstance(marks, (list, tuple)):
        return []
    return [clean(mark) for mark in marks if clean(mark)]


def is_personal_leave_mark(mark: str) -> bool:
    normalized = clean(mark)
    return normalized in {"事", "事假"} or normalized.startswith("事假")


def is_project_transfer_mark(mark: str) -> bool:
    text = clean(mark)
    if not text or text in KNOWN_ATTENDANCE_MARKS or is_personal_leave_mark(text):
        return False
    return any(token in text for token in ("项目", "调入", "调往", "调动")) or (
        text.startswith("在") and "项目" in text
    )


def review_leave_labels(record: dict[str, Any]) -> list[str]:
    labels: list[str] = []
    marks = attendance_mark_values(record)
    for mark in marks:
        if any(token in mark for token in REVIEW_LEAVE_MARKS):
            labels.append(mark)
    for field, caption in (
        ("maternity_leave_days", "产假"),
        ("paternity_leave_days", "陪产假"),
        ("home_leave_days", "探亲假"),
        ("marriage_leave_days", "婚假"),
        ("official_leave_days", "公假"),
        ("study_leave_days", "学习"),
    ):
        if number(record.get(field)):
            labels.append(caption)
    return list(dict.fromkeys(labels))


def month_length(month: str | None) -> int | None:
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", (month or "").strip())
    if not match:
        return None
    return calendar.monthrange(int(match.group(1)), int(match.group(2)))[1]


def attendance_days(record: dict[str, Any], month: str | None = None) -> tuple[float | None, float]:
    actual = number(record.get(
        "actual_work_days",
        record.get(
            "actual_days",
            record.get(
                "actual_attendance",
                record.get(
                    "actualDays",
                    record.get("实际出勤天数", record.get("实际出勤", record.get("出勤天数"))),
                ),
            ),
        ),
    ))
    marks = attendance_mark_values(record)
    if actual is None:
        actual = float(sum(1 for mark in marks if is_attendance_mark(mark)))
        if not marks:
            actual = None
    # 有格子时只数「事/事假」，忽略 JSON 里误填的 personal_leave_days。
    # 无格子时采用 personal_leave_days；leave_days 不是事假。
    # 不得因出勤+事假=当月日历天数而清零（连续事假如 17+13=30）。
    if marks:
        return actual, float(sum(1 for mark in marks if is_personal_leave_mark(mark)))
    explicit = number(record.get(
        "personal_leave_days",
        record.get("personal_leave", record.get("事假天数")),
    ))
    if explicit:
        return actual, float(explicit)
    return actual, 0


def has_attendance_evidence(record: dict[str, Any]) -> bool:
    fields = (
        "actual_work_days", "actual_days", "actual_attendance", "actualDays",
        "实际出勤天数", "实际出勤", "出勤天数",
    )
    return any(number(record.get(field)) is not None for field in fields) or bool(
        record.get("daily_marks") or record.get("days")
    )


def historical_overtime_eligible(base: dict[str, Any]) -> bool:
    """只看历史加班「金额」是否有值。标准列常年有 ROUND 公式，个数/金额为 0 表示当月不发。"""
    if not base:
        return False
    amount = number(base.get("overtime_amount"))
    return amount is not None and amount > 0


def explicit_overtime_days(record: dict[str, Any]) -> float | None:
    """Read explicit overtime totals before deriving them from attendance days."""
    fields = (
        "overtime_days", "weekend_overtime_days", "holiday_overtime_days",
        "加班天数", "周末加班天数", "节假日加班天数",
    )
    values = [number(record.get(field)) for field in fields]
    values = [value for value in values if value is not None]
    if values:
        return sum(values)
    marks = record.get("daily_marks", record.get("days", []))
    if not isinstance(marks, dict):
        return None
    count = 0
    for mark in marks.values():
        normalized = clean(mark)
        if "加班" in normalized or normalized in {"节假日", "周末班"}:
            count += 1
    return float(count) if count else None


def month_number(month: str | None) -> int | None:
    if not month:
        return None
    match = re.search(r"(?:-|/)(\d{1,2})$", month)
    return int(match.group(1)) if match else None


def schedule_defaults(month: str | None, schedule: str | None) -> tuple[float | None, float | None]:
    """Derive normal workdays and payable rest-day overtime capacity."""
    if not month or schedule not in {"single", "double"}:
        return None, None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None, None
    year, month_number_value = int(match.group(1)), int(match.group(2))
    _, days_in_month = calendar.monthrange(year, month_number_value)
    rest_weekdays = {6} if schedule == "single" else {5, 6}
    rest_days = sum(
        1
        for day in range(1, days_in_month + 1)
        if calendar.weekday(year, month_number_value, day) in rest_weekdays
    )
    return float(days_in_month - rest_days), float(rest_days)


def weekend_slots(month: str | None) -> int | None:
    """Return the number of weekend groups in a calendar month."""
    if not month:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None
    year, month_value = int(match.group(1)), int(match.group(2))
    _, days_in_month = calendar.monthrange(year, month_value)
    return len({
        datetime(year, month_value, day).isocalendar()[:2]
        for day in range(1, days_in_month + 1)
        if datetime(year, month_value, day).weekday() in {5, 6}
    })


def calendar_base_workdays(month: str | None, holiday_values: Any) -> float | None:
    """Calculate workdays from month-specific holidays and make-up working days."""
    if not month:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None
    year, month_value = int(match.group(1)), int(match.group(2))
    _, days_in_month = calendar.monthrange(year, month_value)
    official = OFFICIAL_CALENDAR.get(year, {})
    working_weekends = official.get("working_weekends", set())
    non_working = {
        datetime(year, month_value, day).date().isoformat()
        for day in range(1, days_in_month + 1)
        if datetime(year, month_value, day).weekday() in {5, 6}
        and datetime(year, month_value, day).date().isoformat() not in working_weekends
    }
    # Vision results are evidence, not an authority over published calendars.
    # Always retain known statutory days; merge extra supplied dates afterwards.
    official_holidays = official.get("holidays", set())
    supplied_holidays = holiday_values if isinstance(holiday_values, list) else []
    for value in set(official_holidays) | {clean(item)[:10] for item in supplied_holidays}:
        date_text = clean(value)[:10]
        if date_text.startswith(f"{year:04d}-{month_value:02d}-") and re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
            non_working.add(date_text)
    return float(days_in_month - len(non_working))


def month_statutory_holidays(month: str | None, holiday_values: Any) -> list[str]:
    """Return statutory holiday dates in the payroll month, including weekend holidays.

    Ordinary Saturdays and Sundays are not statutory holidays. Vision may still
    supply weekend dates; those are ignored here unless they are on the official
    holiday calendar.
    """
    if not month:
        return []
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return []
    year, month_value = int(match.group(1)), int(match.group(2))
    official = OFFICIAL_CALENDAR.get(year, {}).get("holidays", set())
    supplied = holiday_values if isinstance(holiday_values, list) else []
    prefix = f"{year:04d}-{month_value:02d}-"
    dates: set[str] = set()
    for value in official:
        date_text = clean(value)[:10]
        if date_text.startswith(prefix) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
            dates.add(date_text)
    for value in supplied:
        date_text = clean(value)[:10]
        if not date_text.startswith(prefix) or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_text):
            continue
        if date_text in official:
            dates.add(date_text)
            continue
        try:
            weekday = datetime.strptime(date_text, "%Y-%m-%d").weekday()
        except ValueError:
            continue
        if weekday < 5:
            dates.add(date_text)
    return sorted(dates)


def applicable_phone(
    row: dict[str, Any],
    historical_rows: list[dict[str, Any]],
    month: str | None,
) -> tuple[float, str | None]:
    """Pay a quarterly phone allowance once, in the quarter's final month."""
    month_no = month_number(month)
    if month_no is None:
        return number(row.get("phone")) or 0, None
    if month_no not in (3, 6, 9, 12):
        return 0, None
    quarter_start = month_no - 2
    current_month = month or ""
    current_year = current_month[:4]
    quarter_key = {
        3: "phone_1_3", 6: "phone_4_6",
        9: "phone_7_9", 12: "phone_10_12",
    }[month_no]
    prior_paid = any(
        source_month_value
        and source_month_value[:4] == current_year
        and quarter_start <= int(source_month_value[-2:]) < month_no
        and (number(source.get(quarter_key)) is not None or number(source.get("phone")) is not None)
        for source in historical_rows
        for source_month_value in [source.get("_source_month")]
    )
    if prior_paid:
        return 0, "本季度历史月份已发放话补，当前月份不重复计入"
    amount = number(row.get(quarter_key))
    if amount is not None:
        return amount, None
    fallback = next(
        (
            number(source.get(quarter_key)) or number(source.get("phone"))
            for source in sorted(
                historical_rows,
                key=lambda item: item.get("_source_month") or "",
                reverse=True,
            )
            if number(source.get(quarter_key)) is not None or number(source.get("phone")) is not None
        ),
        None,
    )
    if fallback is None:
        return 0, "历史工资表没有可参考的话补金额"
    return fallback, "话补金额沿用最近历史记录，需人工复核"


def is_attendance_mark(mark: Any) -> bool:
    return clean(mark).lower() in {
        "8", "出勤", "上班", "正常", "加班", "周末加班", "节假日加班",
        "√", "✔", "1",
    }


def known_attendance_mark(mark: Any) -> bool:
    text = clean(mark)
    if text in KNOWN_ATTENDANCE_MARKS or is_attendance_mark(text) or is_personal_leave_mark(text):
        return True
    if is_project_transfer_mark(text):
        return True
    return any(token in text for token in REVIEW_LEAVE_MARKS)


def counted_attendance_marks(record: dict[str, Any], month: str | None = None) -> float | None:
    """Count 8/√ attendance marks when the grid is complete enough to audit."""
    explicit = number(record.get("marked_work_days", record.get("符号出勤天数")))
    if explicit is not None:
        return explicit
    raw = record.get("daily_marks", record.get("days"))
    needed = month_length(month) or 20
    if isinstance(raw, dict):
        values = list(raw.values())
        if len(raw) < needed:
            return None
    elif isinstance(raw, (list, tuple)):
        values = list(raw)
        if sum(1 for value in values if clean(value)) < needed:
            return None
    else:
        return None
    return float(sum(1 for mark in values if is_attendance_mark(mark)))


def attendance_mark_mismatch_issue(
    record: dict[str, Any],
    actual: float | None,
    month: str | None,
) -> str | None:
    """Require printed 实际出勤 to equal the count of attendance marks."""
    counted = counted_attendance_marks(record, month)
    if actual is None or counted is None or math.isclose(actual, counted, abs_tol=1e-6):
        return None
    printed = int(actual) if float(actual).is_integer() else actual
    marks = int(counted) if float(counted).is_integer() else counted
    return (
        f"印刷实际出勤与出勤符号天数不一致（印刷{printed}，符号{marks}），"
        "工资按印刷实际出勤计算，需人工复核"
    )


def attendance_overtime_days(
    record: dict[str, Any],
    month: str | None,
    work_schedule: str | None,
) -> float | None:
    """Count actual attendance on scheduled rest days and named holidays."""
    marks = record.get("daily_marks", record.get("days"))
    weekend_values = record.get("weekend_attendance_dates")
    holiday_attendance_values = record.get("holiday_attendance_dates")
    has_date_evidence = isinstance(marks, dict) or isinstance(weekend_values, list) or isinstance(holiday_attendance_values, list)
    if not has_date_evidence or not month:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{1,2})", month.strip())
    if not match:
        return None
    year, month_value = int(match.group(1)), int(match.group(2))
    holiday_values = record.get("holiday_dates", record.get("holidays", []))
    holiday_dates = {clean(value)[:10] for value in holiday_values} if isinstance(holiday_values, list) else set()
    weekend_dates = {clean(value)[:10] for value in weekend_values} if isinstance(weekend_values, list) else set()
    holiday_attendance_dates = (
        {clean(value)[:10] for value in holiday_attendance_values}
        if isinstance(holiday_attendance_values, list) else set()
    )
    weekend_by_week: defaultdict[tuple[int, int], int] = defaultdict(int)
    holiday_count = 0
    attendance_dates = set(weekend_dates) | set(holiday_attendance_dates)
    mark_items = marks.items() if isinstance(marks, dict) else []
    for raw_date, mark in mark_items:
        if is_attendance_mark(mark):
            attendance_dates.add(clean(raw_date)[:10])
    for date_text in attendance_dates:
        try:
            date = datetime.strptime(date_text, "%Y-%m-%d")
        except ValueError:
            continue
        if date.year != year or date.month != month_value:
            continue
        if date.weekday() in {5, 6}:
            weekend_by_week[date.isocalendar()[:2]] += 1
        elif date_text in holiday_dates or date_text in holiday_attendance_dates:
            holiday_count += 1
    # A project weekend is one calendar Saturday/Sunday group. Working either
    # Saturday or Sunday earns one overtime day; the schedule picker is not
    # needed to decide this.
    # Each worked weekend date counts. The monthly weekend-slot cap is applied
    # by calculate(), so working both Saturday and Sunday contributes two
    # days before the cap is enforced.
    weekend_overtime = sum(weekend_by_week.values())
    return float(weekend_overtime + holiday_count)


def group_attendance_records(attendance: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    """Merge same-person attendance across projects; split only by vendor category."""
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in attendance:
        identity = row_identity(record)
        if identity[0]:
            by_name[identity[0]].append(record)
    grouped: list[list[dict[str, Any]]] = []
    for records in by_name.values():
        buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        unspecified: list[dict[str, Any]] = []
        for record in records:
            category = clean(record.get("category"))
            if is_employment_category(category) and not looks_like_department_or_title(category):
                buckets[category].append(record)
            else:
                unspecified.append(record)
        if not buckets:
            grouped.append(records)
            continue
        if len(buckets) == 1:
            grouped.append(next(iter(buckets.values())) + unspecified)
            continue
        for category, items in buckets.items():
            grouped.append(items)
        if unspecified:
            grouped.append(unspecified)
    return grouped


def attendance_segments(records: list[dict[str, Any]], month: str | None) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    for record in records:
        explicit = record.get("project_segments")
        if isinstance(explicit, list) and explicit:
            for item in explicit:
                if not isinstance(item, dict):
                    continue
                days = number(item.get("actual_work_days", item.get("days")))
                project = content_project(item) or clean(item.get("project"))
                segments.append({"project": project, "actual_work_days": days, "record": record})
            continue
        days, _ = attendance_days(record, month)
        project = content_project(record)
        segments.append({"project": project, "actual_work_days": days, "record": record})
    merged: dict[str, dict[str, Any]] = {}
    for segment in segments:
        key = segment["project"] or ""
        current = merged.get(key)
        if current is None:
            merged[key] = dict(segment)
            continue
        left, right = current.get("actual_work_days"), segment.get("actual_work_days")
        current["actual_work_days"] = (left or 0) + (right or 0) if left is not None or right is not None else None
    return list(merged.values())


def display_project(segments: list[dict[str, Any]], history_row: dict[str, Any]) -> str:
    ranked = sorted(
        (segment for segment in segments if segment.get("project") and segment.get("actual_work_days")),
        key=lambda item: item.get("actual_work_days") or 0,
        reverse=True,
    )
    if ranked:
        return ranked[0]["project"]
    for segment in segments:
        if segment.get("project"):
            return segment["project"]
    return content_project(history_row)


def colleague_construction_rate(historical_rows: list[dict[str, Any]], project: str, exclude_name: str) -> float | None:
    if not project:
        return None
    rates: set[float] = set()
    zeros = 0
    seen = 0
    for row in historical_rows:
        if normalized_name(row.get("name")) == exclude_name:
            continue
        row_project = content_project(row)
        if row_project and not projects_compatible(project, row_project):
            continue
        if row_project and project and not projects_compatible(project, row_project):
            continue
        if not row_project:
            continue
        terms = parse_construction_formula(row.get("_construction_formula"))
        if len(terms) == 1 and terms[0][0] is None:
            seen += 1
            rates.add(terms[0][1])
            continue
        raw = number(row.get("construction"))
        if raw == 0:
            zeros += 1
            seen += 1
    if len(rates) == 1:
        return next(iter(rates))
    if seen and zeros == seen:
        return 0.0
    return None


def resolve_construction(
    base: dict[str, Any],
    segments: list[dict[str, Any]],
    historical_rows: list[dict[str, Any]],
    actual: float | None,
) -> tuple[float, float | None, list[tuple[float, float]], list[str]]:
    issues: list[str] = []
    formula = base.get("_construction_formula")
    terms = parse_construction_formula(formula)
    raw = number(base.get("construction"))
    explicit_day = number(base.get("construction_day"))
    historical_workdays = number(base.get("work_days"))
    name = normalized_name(base.get("name"))
    cross_project = len([item for item in segments if item.get("project")]) > 1 or len(terms) > 1

    if explicit_day is not None:
        amount = (actual or 0) * explicit_day
        return amount, explicit_day, [((actual or 0), explicit_day)], issues

    if terms:
        if construction_formula_zeroed(formula):
            rate = terms[0][1]
            return 0.0, rate, [], issues
        if len(terms) == 1 and terms[0][0] is None:
            rate = terms[0][1]
            return (actual or 0) * rate, rate, [((actual or 0), rate)], issues
        if len(terms) == 1 and terms[0][0] is not None:
            rate = terms[0][1]
            return (actual or 0) * rate, rate, [((actual or 0), rate)], issues
        excel_terms: list[tuple[float, float]] = []
        remaining = [dict(item) for item in segments if item.get("actual_work_days") is not None]
        for days, rate in terms:
            if days is None:
                excel_terms.append(((actual or 0), rate))
                continue
            match = next((item for item in remaining if math.isclose(item.get("actual_work_days") or 0, days, abs_tol=1e-6)), None)
            if match:
                remaining.remove(match)
            excel_terms.append((days, rate))
        amount = sum(days * rate for days, rate in excel_terms)
        unique_rates = {rate for _, rate in excel_terms}
        day_rate = next(iter(unique_rates)) if len(unique_rates) == 1 else None
        return amount, day_rate, excel_terms, issues

    if raw == 0:
        return 0.0, 0.0, [((actual or 0), 0.0)], issues

    if raw is not None and historical_workdays and historical_workdays > 0 and not cross_project:
        rate = raw / historical_workdays
        issues.append("施工补贴沿用历史工资基准，需人工复核")
        return (actual or 0) * rate, rate, [((actual or 0), rate)], issues

    if raw is not None and not historical_workdays:
        issues.append("施工补贴只有月金额且缺少历史工作天数，未猜测日标准")
        return 0.0, None, [], issues

    if cross_project:
        excel_terms = []
        unresolved = False
        for segment in segments:
            project = segment.get("project") or ""
            days = segment.get("actual_work_days") or 0
            rate = colleague_construction_rate(historical_rows, project, name)
            if rate is None:
                unresolved = True
                continue
            excel_terms.append((days, rate))
        if unresolved or not excel_terms:
            issues.append("缺少施工补贴日标准")
            return 0.0, None, [], issues
        amount = sum(days * rate for days, rate in excel_terms)
        unique_rates = {rate for _, rate in excel_terms}
        day_rate = next(iter(unique_rates)) if len(unique_rates) == 1 else None
        issues.append("施工补贴日标准由同项目同事公式推导，需人工复核")
        return amount, day_rate, excel_terms, issues

    issues.append("缺少施工补贴日标准")
    return 0.0, None, [], issues


def choose_history_baseline(
    record: dict[str, Any],
    identity: tuple[str, str, str],
    historical_rows: list[dict[str, Any]],
    month: str | None,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    """Select an auditable baseline without silently discarding duplicate history."""
    issues: list[str] = []
    matches = [row for row in historical_rows if identity_matches(identity, row_identity(row))]
    if not matches:
        fuzzy, reason, ambiguity = uniquely_fuzzy_matched(record, historical_rows, identity)
        if ambiguity:
            return {}, [], [ambiguity]
        if fuzzy is None:
            return {}, [], []
        matches = [
            row for row in historical_rows
            if identity_matches(row_identity(fuzzy), row_identity(row))
        ]
        issues.append(fuzzy_match_review_note(reason))
    same_month = [row for row in matches if row.get("_source_month") == month]
    candidates = same_month or matches
    dated = [row for row in candidates if row.get("_source_month")]
    if dated:
        newest = max(row["_source_month"] for row in dated)
        candidates = [row for row in candidates if row.get("_source_month") == newest]
    if len(candidates) != 1:
        sources = "、".join(
            f"{row.get('_source_file', row.get('_sheet', '未知来源'))}!{row.get('_row', '')}"
            for row in candidates[:3]
        )
        return {}, matches, [f"历史工资存在多个同月候选，无法唯一匹配：{sources}"]
    return candidates[0], matches, issues


def calculate(
    roster_rows: list[dict[str, Any]],
    historical_rows: list[dict[str, Any]],
    attendance: list[dict[str, Any]],
    month: str | None,
    base_workdays: float | None,
    overtime_cap: float | None,
    work_schedule: str | None = None,
    attendance_project: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    derived_workdays, derived_overtime_cap = schedule_defaults(month, work_schedule)
    base_workdays = base_workdays if base_workdays is not None else derived_workdays
    overtime_cap = overtime_cap if overtime_cap is not None else derived_overtime_cap
    if attendance_project:
        stamped = []
        for record in attendance:
            if not isinstance(record, dict):
                continue
            if not content_project(record):
                stamped.append({**record, "page_project": attendance_project, "project": record.get("project") or attendance_project})
            else:
                stamped.append(record)
        attendance = stamped
    roster: dict[tuple[str, str, str], dict[str, Any]] = {}
    ignored = []
    for row in roster_rows:
        identity = row_identity(row)
        if not identity[0]:
            continue
        if "在册" in clean(row.get("status")) or "在册" in clean(row.get("category")):
            ignored.append({**row, "reason": "在册人员跳过核算"})
        else:
            roster[identity] = row

    historical_rows = [row for row in historical_rows if row_identity(row)[0]]
    month_no = month_number(month)
    details, baselines, attendance_out, exceptions, reconciliation = [], [], [], [], []
    for records in group_attendance_records(attendance):
        record = records[0]
        identity = row_identity(record)
        name = identity[0]
        ignored_match = next(
            (row for row in ignored if identity_matches(identity, row_identity(row))),
            None,
        )
        if ignored_match is not None:
            continue
        issues: list[str] = []
        roster_row = roster.get(identity, {})
        if not roster_row:
            roster_candidates = [
                row for roster_identity, row in roster.items()
                if identity_matches(identity, roster_identity)
            ]
            if len(roster_candidates) == 1:
                roster_row = roster_candidates[0]
            elif not roster_candidates:
                roster_row, reason, ambiguity = uniquely_fuzzy_matched(
                    observed=record, candidates=list(roster.values()),
                )
                if roster_row is not None:
                    issues.append(fuzzy_match_review_note(reason))
                elif ambiguity:
                    issues.append(ambiguity)
                roster_row = roster_row or {}
            elif len(roster_candidates) > 1:
                issues.append("考勤人员未在外聘人员底表唯一匹配")
        if not roster_row:
            issues.append("考勤人员未在外聘人员底表唯一匹配")
        baseline_identity = row_identity(roster_row) if roster_row else identity
        base, matched_history, history_issues = choose_history_baseline(
            record, baseline_identity, historical_rows, month,
        )
        issues.extend(history_issues)
        segments = attendance_segments(records, month)
        resolved_project = display_project(segments, base)
        resolved_category = (
            clean(roster_row.get("category")) or clean(base.get("category"))
        )
        if looks_like_department_or_title(resolved_category):
            resolved_category = ""
        if not base:
            exceptions.append({
                "姓名": record.get("name") or name, "项目": resolved_project,
                "人员类别": resolved_category, "类型": "需复核",
                "说明": "；".join(dict.fromkeys(issues + ["未匹配历史工资，未计算"])),
            })
            continue
        actual_parts = [item.get("actual_work_days") for item in segments]
        if any(part is not None for part in actual_parts):
            actual = sum(part or 0 for part in actual_parts)
        else:
            actual = None
        days_in_month = month_length(month)
        if actual is not None and days_in_month is not None and actual > days_in_month:
            issues.append(
                f"合并后实际出勤{actual:g}天超过{month}日历{days_in_month}天，可能跨页重复计数，需核对调动分段"
            )
        hist_days = number(base.get("work_days"))
        if (
            actual is not None
            and hist_days is not None
            and base.get("_source_month") == month
            and not math.isclose(actual, hist_days, abs_tol=1e-6)
        ):
            issues.append(
                f"同月历史工资表工作天数为{hist_days:g}，与考勤实际出勤{actual:g}不一致，请核对本月工资表与考勤"
            )
        personal_leave = 0.0
        sick = 0.0
        absent = 0.0
        for item in records:
            _, leave = attendance_days(item, month)
            personal_leave += leave
            sick += number(item.get("sick_leave_days")) or 0
            absent += number(item.get("absent_days")) or 0
        if not any(has_attendance_evidence(item) for item in records):
            issues.append("缺少实际出勤字段或每日考勤，未将缺失误作零出勤")
            actual = None
        for item in records:
            item_actual, _ = attendance_days(item, month)
            mismatch = attendance_mark_mismatch_issue(item, item_actual, month)
            if mismatch:
                issues.append(mismatch)
        for label, value in (("病假", sick), ("旷工", absent)):
            if value:
                issues.append(f"{label}暂不计算，需人工复核")
        for item in records:
            for mark in review_leave_labels(item):
                issues.append(f"{mark}暂不计算，需人工复核")
            unrecognized = [
                mark for mark in item.get("unrecognized_marks", [])
                if not known_attendance_mark(mark)
            ]
            if unrecognized:
                issues.append("存在未识别考勤符号")
            if item.get("entry_exit_days"):
                issues.append("入离职分项暂不计算")
            for mark in attendance_mark_values(item):
                if not known_attendance_mark(mark):
                    issues.append("存在未识别考勤符号")
                    break

        position = number(base.get("position_salary"))
        performance = number(base.get("performance"))
        seniority = number(base.get("seniority")) or 0
        title = number(base.get("title")) or 0
        transport = number(base.get("transport")) or 0
        construction, construction_day, construction_terms, construction_issues = resolve_construction(
            base, segments, historical_rows, actual,
        )
        issues.extend(construction_issues)
        phone, phone_issue = applicable_phone(base, matched_history, month)
        if phone_issue:
            issues.append(phone_issue)
        hot = (number(base.get("hot")) or 0) if month_no in (6, 7, 8, 9) else 0
        if position is None:
            issues.append("缺少岗位工资/基本工资")
        if performance is None:
            issues.append("缺少绩效工资标准")
        if number(base.get("seniority")) is None:
            issues.append("缺少工龄工资基准")
        adjusted_performance = max(performance - performance / 21.75 * personal_leave, 0) if performance is not None else 0

        pays_overtime = historical_overtime_eligible(base)
        overtime_standard = number(base.get("overtime_standard")) if pays_overtime else None
        if pays_overtime and overtime_standard is None and position is not None:
            overtime_standard = round(position / 21.75) * 2
        overtime_days = 0
        overtime_amount = 0
        effective_base_workdays = calendar_base_workdays(month, record.get("holiday_dates"))
        calendar_overtime_cap = weekend_slots(month)
        effective_overtime_cap = float(calendar_overtime_cap) if calendar_overtime_cap is not None else overtime_cap
        if not pays_overtime:
            overtime_standard = None
        elif effective_base_workdays is None or effective_overtime_cap is None:
            issues.append("缺少当月基础工作日或计划加班上限，加班费未计算")
        elif overtime_standard is not None and actual is not None:
            overtime_days = min(max(actual - effective_base_workdays, 0), effective_overtime_cap)
            overtime_amount = overtime_days * overtime_standard
        total = sum((position or 0, adjusted_performance, seniority, title, construction,
                     phone, hot, transport, overtime_amount))
        display_name = record.get("name") or name
        for field, value in (("sick_leave_days", sick), ("absent_days", absent)):
            if value:
                exceptions.append({"姓名": display_name, "项目": resolved_project, "人员类别": resolved_category,
                                    "类型": "病假" if field == "sick_leave_days" else "旷工",
                                    "说明": "规则未确认，金额未扣除"})
        for issue in dict.fromkeys(issues):
            if any(item.get("姓名") == display_name and item.get("说明") == issue for item in exceptions):
                continue
            exceptions.append({"姓名": display_name, "项目": resolved_project, "人员类别": resolved_category,
                                "类型": "需复核", "说明": issue})
        source = f"历史工资:{base.get('_sheet', '')}!{base.get('_row', '')}" if base.get("_sheet") or base.get("_row") else "历史工资"
        baselines.append({"姓名": display_name, "项目": resolved_project, "人员类别": resolved_category, "来源": source,
                          "岗位工资": position, "绩效工资": performance, "工龄工资": seniority,
                          "职称工资": title, "施工补贴日标准": construction_day,
                          "话费补贴": phone, "高温补贴": hot, "交通补贴": transport,
                          "加班标准": overtime_standard})
        attendance_out.append({"姓名": display_name, "项目": resolved_project, "人员类别": resolved_category,
                               "实际出勤": actual, "事假天数": personal_leave,
                               "病假天数": sick, "旷工天数": absent,
                               "来源": "当月考勤表"})
        details.append({
            "姓名": display_name, "项目": resolved_project, "人员类别": resolved_category, "实际出勤": actual,
            "事假天数": personal_leave, "岗位工资": position or 0,
            "绩效工资": adjusted_performance, "工龄工资": seniority,
            "职称工资": title, "施工补贴": construction, "话费补贴": phone,
            "高温补贴": hot, "交通补贴": transport, "加班天数": overtime_days,
            "加班费": overtime_amount, "应发工资": total,
            "计算状态": "需复核" if issues else "已计算",
            "复核说明": "；".join(issues),
            "_performance_standard": performance,
            "_construction_terms": construction_terms,
            "_pays_overtime": pays_overtime,
            "_overtime_standard": overtime_standard,
        })
        comparable_history = matched_history
        historical_gross = (
            number(comparable_history[0].get("historical_gross"))
            if len(comparable_history) == 1 else None
        )
        comparison_status = (
            "未匹配历史应发" if not comparable_history else
            "历史应发不唯一" if len(comparable_history) > 1 else
            "未匹配历史应发" if historical_gross is None else
            "一致" if abs(total - historical_gross) <= 0.01 else "有差异"
        )
        reconciliation.append({"姓名": display_name, "项目": resolved_project, "人员类别": resolved_category,
                               "计算应发": total, "历史应发": historical_gross,
                               "差异": total - historical_gross if historical_gross is not None else None,
                               "状态": comparison_status})
    for row in ignored:
        exceptions.append({"姓名": row.get("name"), "项目": content_project(row),
                           "人员类别": "在册", "类型": "可忽略",
                           "说明": "在册人员不进入外聘工资核算"})
    return {
        "payroll_detail": grouped_by_project(details),
        "baseline": grouped_by_project(baselines),
        "attendance": grouped_by_project(attendance_out),
        "reconciliation": grouped_by_project(reconciliation),
        "review_exceptions": grouped_by_project(exceptions),
    }


def grouped_by_project(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Keep each project as a contiguous block, in the order projects first appear."""
    project_order: dict[str, int] = {}
    for row in rows:
        key = clean(row.get("项目"))
        if key not in project_order:
            project_order[key] = len(project_order)
    return [
        row for _, row in sorted(
            enumerate(rows),
            key=lambda item: (project_order[clean(item[1].get("项目"))], item[0]),
        )
    ]


def excel_letter(index: int) -> str:
    letter = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter


def visible_headers(row: dict[str, Any]) -> list[str]:
    return [key for key in row if not str(key).startswith("_")]


def payroll_detail_cell(
    header: str,
    row: dict[str, Any],
    excel_row: int,
    columns: dict[str, str],
) -> Any:
    if header == "绩效工资":
        standard = number(row.get("_performance_standard"))
        leave_cell = columns.get("事假天数")
        if standard is None or not leave_cell:
            return row.get(header)
        return f"=MAX({standard}-{standard}/21.75*{leave_cell}{excel_row},0)"
    if header == "施工补贴":
        terms = row.get("_construction_terms") or []
        actual_cell = columns.get("实际出勤")
        if len(terms) == 1 and actual_cell and math.isclose(terms[0][0], number(row.get("实际出勤")) or 0, abs_tol=1e-6):
            return f"={actual_cell}{excel_row}*{terms[0][1]}"
        if terms:
            return "=" + "+".join(f"{days}*{rate}" for days, rate in terms)
        return row.get(header) if row.get(header) else 0
    if header == "加班费":
        if not row.get("_pays_overtime") or number(row.get("_overtime_standard")) is None:
            return row.get(header) if row.get(header) else 0
        days_cell = columns.get("加班天数")
        return f"={days_cell}{excel_row}*{row.get('_overtime_standard')}"
    if header == "应发工资":
        parts = [
            columns[name] + str(excel_row)
            for name in (
                "岗位工资", "绩效工资", "工龄工资", "职称工资", "施工补贴",
                "话费补贴", "高温补贴", "交通补贴", "加班费",
            )
            if name in columns
        ]
        return "=" + "+".join(parts) if parts else row.get(header)
    return row.get(header)


def write_workbook(result: dict[str, list[dict[str, Any]]], output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sheet_titles = {
        "payroll_detail": GENERATED_PAYROLL_SHEETS[0],
        "baseline": GENERATED_PAYROLL_SHEETS[1],
        "attendance": GENERATED_PAYROLL_SHEETS[2],
        "reconciliation": GENERATED_PAYROLL_SHEETS[3],
        "review_exceptions": GENERATED_PAYROLL_SHEETS[4],
    }
    for sheet_name, title in sheet_titles.items():
        sheet = workbook.create_sheet(title)
        sheet.freeze_panes = "A2"
        rows = result[sheet_name]
        headers = visible_headers(rows[0]) if rows else ["说明"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        columns = {header: excel_letter(index) for index, header in enumerate(headers, 1)}
        for offset, row in enumerate(rows, 2):
            values = [
                payroll_detail_cell(header, row, offset, columns)
                if sheet_name == "payroll_detail" else row.get(header)
                for header in headers
            ]
            sheet.append(values)
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in column) + 2, 10), 32)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00;(#,##0.00);0.00'
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser(description="外聘工资确定性计算器（本地文件、无网络）")
    result.add_argument("--roster", required=True, nargs="+", type=Path, help="一个或多个人员底表 xlsx")
    result.add_argument("--historical", required=True, nargs="+", type=Path, help="一个或多个历史工资 xlsx")
    result.add_argument("--attendance", required=True, nargs="+", type=Path, help="一个或多个视觉模型输出的考勤 JSON")
    result.add_argument("--output", required=True, type=Path, help="输出工作簿 xlsx")
    result.add_argument("--month", help="核算月份，例如 2026-06；也可由 attendance.month 提供")
    result.add_argument("--base-workdays", type=float, help="当月基础工作日，缺少时可由工休制度推导")
    result.add_argument("--overtime-cap", type=float, help="当月计划加班上限，缺少时可由工休制度推导")
    result.add_argument("--work-schedule", choices=("single", "double"), help="项目工休制度：single 单休，double 双休")
    return result


def main(argv: list[str] | None = None) -> int:
    args = parser().parse_args(argv)
    try:
        for path in [*args.roster, *args.attendance, *args.historical]:
            if not path.is_file():
                raise ValueError(f"输入文件不存在：{path}")
        att_month, attendance_project, records = attendance_records_from_paths(list(args.attendance))
        month = args.month or att_month
        holiday_values = []
        for record in records:
            if isinstance(record, dict) and isinstance(record.get("holiday_dates"), list):
                holiday_values.extend(record.get("holiday_dates"))
        historical_rows, skipped_generated = load_historical_payroll(list(args.historical))
        result = calculate(
            [row for path in args.roster for row in read_rows(path)],
            historical_rows,
            records, month, args.base_workdays, args.overtime_cap,
            args.work_schedule, attendance_project,
        )
        if not records:
            raise ValueError("没有任何可识别的考勤人员")
        input_notes = note_skipped_generated_history(result, skipped_generated)
        write_workbook(result, args.output)
        print(json.dumps({"status": "success", "output": str(args.output),
                          "month": month,
                          "base_workdays": calendar_base_workdays(month, holiday_values),
                          "overtime_cap": weekend_slots(month),
                          "statutory_holidays": month_statutory_holidays(month, holiday_values),
                          "rows": {name: len(rows) for name, rows in result.items()},
                          "review_count": len(result["review_exceptions"]),
                          **input_notes},
                         ensure_ascii=False, sort_keys=True))
        return 0
    except (ValueError, OSError, json.JSONDecodeError) as exc:
        print(json.dumps({"status": "error", "code": ERROR_BLOCKED if "考勤" in str(exc) else ERROR_INPUT,
                          "error": str(exc)}, ensure_ascii=False, sort_keys=True))
        return ERROR_BLOCKED if "考勤" in str(exc) else ERROR_INPUT
    except Exception as exc:  # keep CLI errors stable without leaking local contents
        print(json.dumps({"status": "error", "code": ERROR_OUTPUT, "error": type(exc).__name__},
                         ensure_ascii=False, sort_keys=True))
        return ERROR_OUTPUT


if __name__ == "__main__":
    sys.exit(main())
